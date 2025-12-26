#OLÁ!
#PROCEDIMENTO: SOLICITAR REPASSE FINANCEIRO;
#POR: LOUISE-SESDEC;
#ALTERAÇÕES NO CÓDIGO PODEM SER ACESSADAS NO MEU GITHUB: <https://github.com/louisegoncalves/Automatizacao-SIGEF>.

#INSTRUÇÕES
#ATENÇÃO: É OBRIGATÓRIO ABRIR O DEPURADOR DO GOOGLE CHROME PARA EXECUTAR ESSE CÓDIGO
#EXECUTE NO CMD: "C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebugProfile"
#E LOGUE NO SIGEF


ug = '150001'
gestao = '00001'
#BIBLIOTECAS UTILIZADAS:
from playwright.sync_api import sync_playwright, Page, TimeoutError
import pyautogui
import openpyxl
import keyboard
import time
import pyperclip
import re
from datetime import date
from datetime import datetime
import sys

#QUAL PLANILHA VAI SER UTILIZADA?
#planilha = "Pagamentos - voluntariar novembro.xlsx"
planilha = "Pagamentos.xlsx"

#VARIÁVEIS IMPORTANTES:
robo_deve_parar = False
coluna = 1
linha = 2
loop = True
despesa_certificada_teste = 'None'

#PLANILHA NO EXCEL:
try:
    book = openpyxl.load_workbook(planilha)
    pagina = book['Entrada']
    pagina1 = book['Despesas Certificadas']
    pagina2 = book['Notas de Liquidação']
    pagina3 = book['Preparações de Pagamento']
    pagina4 = book['Ordens Bancárias']
    pagina5 = book['Saída']
except: 
    pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
    sys.exit()

#PLANILHA DE BACKUP:
try:
    book1 = openpyxl.load_workbook("Pagamentos_Backup.xlsx")
    pagina_backup = book1['Entrada']
    pagina1_backup = book1['Despesas Certificadas']
    pagina2_backup = book1['Notas de Liquidação']
    pagina3_backup = book1['Preparações de Pagamento']
    pagina4_backup = book1['Ordens Bancárias']
    pagina5_backup = book1['Saída']
except:
    try:
        wb = openpyxl.Workbook()
        ws_principal = wb.active
        ws_principal.title = "Entrada"
        wb.save("Pagamentos_Backup.xlsx")
        try:
            wb.create_sheet("Despesas Certificadas")
            wb.create_sheet("Notas de Liquidação")
            wb.create_sheet("Preparações de Pagamento")
            wb.create_sheet("Ordens Bancárias")
            wb.create_sheet('Saída')
            wb.save("Pagamentos_Backup.xlsx")
            print("Arquivo 'Pagamentos_Backup.xlsx' criado com sucesso com várias planilhas.")
        except:
            print("Erro na planilha de backup.")
        try:
            book1 = openpyxl.load_workbook("Pagamentos_Backup.xlsx")
            pagina_backup = book1['Entrada']
            pagina1_backup = book1['Despesas Certificadas']
            pagina2_backup = book1['Notas de Liquidação']
            pagina3_backup = book1['Preparações de Pagamento']
            pagina4_backup = book1['Ordens Bancárias']
            pagina5_backup = book1['Saída']
        except:
                print("Erro na planilha de backup.")
    except:
        print("Erro na planilha de backup.")

#SE QUISER DESATIVAR AQUELA JANELA DO COMEÇO PODE EXCLUIR ELA AQUI:
pyautogui.alert(text='Procedimento: Solicitar Repasse Financeiro', title='Início', button='OK')

#FUNÇÃO QUE SERÁ CHAMADA PELA TECLA DE PÂNICO:
def parar_execucao():
    global robo_deve_parar
    print("\n!!! TECLA ESC ACIONADA! ENCERRANDO AUTOMACAO !!!")
    robo_deve_parar = True

#FUNÇÃO QUE ENCERRA O CODIGO E FECHA A PLANILHA COM SEGURANÇA:
#A PLANILHA DEVE SEMPRE SER FECHADA ANTES DE ENCERRAR, POIS CORRE O RISCO DE CORROMPER!
def verificar_panico_e_sair(workbook):
    global robo_deve_parar
    if robo_deve_parar:
        print("Garantindo que a planilha seja fechada...")
        if workbook:
            workbook.close()
        pyautogui.alert('Tecla ESC acionada. Automação encerrada.')
        sys.exit()

#DEFINA SUA TECLA DE PÂNICO:
tecla_de_panico = "Esc" 
keyboard.add_hotkey(tecla_de_panico, parar_execucao)
print(f"Robô iniciado. Pressione a tecla '{tecla_de_panico}' a qualquer momento para abortar com seguranca.")

#AQUI ELE VAI PEDIR PARA ABRIR O SIGEF PELO DEPURADOR DO GOOGLE:
pyautogui.confirm(text='Aperte OK quando o SIGEF estiver logado no depurador do Google Chrome', title='Depurador do Chrome' , buttons=['OK'])

#PORTA DO DEPURADOR DO GOOGLE CHROME
CHROME_DEBUG_URL = "http://localhost:9222"

if robo_deve_parar:
    verificar_panico_e_sair(book)

#EXECUTANDO O PLAYWRIGHT DE FORMA SÍNCRONA:
with sync_playwright() as p:
        if robo_deve_parar:
            verificar_panico_e_sair(book)

        #CONECTAR AO NAVEGADOR JÁ ABERTO:
        print(f"Tentando se conectar ao Chrome na porta de depuração: {CHROME_DEBUG_URL}")
        browser = p.chromium.connect_over_cdp(CHROME_DEBUG_URL)
        print("Conexão estabelecida com sucesso!")

        #OBTER A PÁGINA QUE ESTÁ ABERTA:
        #Quando conectamos, precisamos pegar o contexto e a página existentes;
        #Geralmente, a primeira página do primeiro contexto é a que queremos.
        janela = browser.contexts[0]
        guia = janela.pages[0]

        #VERIFICAR A PÁGINA ABERTA:
        print(f"Assumindo o controle da página com o título: '{guia.title()}'")
            
        #LOCALIZANDO O IFRAME:
        frame = guia.frame_locator('iframe[src="/SIGEF2025/SEG/#/SEGControleAcesso?p=1"]')
        
        #VERIFICAÇÃO DE PÂNICO:
        if robo_deve_parar:
            verificar_panico_e_sair(book)
        
        #INÍCIO                
        print("Iniciando!")

        #PESQUISANDO FUNCIONALIDADE NO SIGEF:
        pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
        pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
        pesquisar_funcionalidades_sistema.press_sequentially("Solicitar Repasse Financeiro")
        funcionalidade_sistema = frame.get_by_title("Solicitar Repasse Financeiro", exact=True)
        
        #ABRINDO JANELA SOLICITAR REPASSE FINANCEIRO:
        with guia.expect_popup() as popup_info:
            funcionalidade_sistema.click()
            Solicitar_Repasse_Financeiro = popup_info.value

            Preencher_Unidade_Gestora = Solicitar_Repasse_Financeiro.locator("#txtUnidadeGestora")
            Preencher_Gestao = Solicitar_Repasse_Financeiro.locator("#txtGestao_SIGEFPesquisa")
            Preencher_Grupo_Financeiro = Solicitar_Repasse_Financeiro.locator("#txtGrupo_SIGEFPesquisa")
            Botão_Pesquisar = Solicitar_Repasse_Financeiro.get_by_role("button", name="Confirmar a Pesquisa")



            Preencher_Unidade_Gestora.press("Control+KeyA+Backspace")
            Preencher_Unidade_Gestora.press_sequentially(ug)
            Preencher_Gestao.press_sequentially(gestao)
            Grupo_Financeiro = pyautogui.confirm(text='Escolha o Grupo Financeiro', title='Grupo Financeiro' , buttons=['393', '314','315'])
            Preencher_Grupo_Financeiro.press_sequentially(Grupo_Financeiro)
            Botão_Pesquisar.click()
            Grupo_Financeiro_Encontrado = Solicitar_Repasse_Financeiro.get_by_text(Grupo_Financeiro)
            Grupo_Financeiro_Encontrado.wait_for()
            Grupo_Financeiro_Encontrado.click()
            with Solicitar_Repasse_Financeiro.expect_popup() as popup_info:
                Solicitar_Repasse_Financeiro_Interno = popup_info.value
                Setinha_A_Solicitar = Solicitar_Repasse_Financeiro_Interno.locator("[id=\"dtgLancamentos_ctl03_lnkControle_'gifAguarde'); CopiarValor(4,1); window.setTimeout('EscondeImgEsperar(\"]")
                Setinha_A_Solicitar.wait_for()
                Campo_Justificativa = Solicitar_Repasse_Financeiro_Interno.locator("#txtJustificativa")
                Botao_Confirmar = Solicitar_Repasse_Financeiro_Interno.get_by_role("button", name="Confirmar a Operação")
                Setinha_A_Solicitar.click()
                Campo_Justificativa.press_sequentially("Solicitação de Repasse Financeiro.")
                Botao_Confirmar.click()

                sys.exit()


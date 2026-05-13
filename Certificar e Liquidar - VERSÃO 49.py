# =================================================================
#OLÁ!
#PROCEDIMENTO: CERTIFICAR E LIQUIDAR;
#VERSÃO 49 EM 13/05/2026;
#POR: LOUISE-SESDEC;
#ALTERAÇÕES NO CÓDIGO PODEM SER ACESSADAS NO MEU GITHUB: <https://github.com/louisegoncalves/Automatizacao-SIGEF>.
# =================================================================

# =================================================================
#INSTRUÇÕES:
#ATENÇÃO: É OBRIGATÓRIO ABRIR O DEPURADOR DO GOOGLE CHROME PARA EXECUTAR ESSE CÓDIGO;
#EXECUTE NO CMD: ""C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebugProfile"
#E LOGUE NO SIGEF.
# =================================================================

# =================================================================
#BIBLIOTECAS UTILIZADAS:
from playwright.sync_api import sync_playwright
import pyautogui
import openpyxl
import os
import keyboard
import time
import pyperclip
import re
from datetime import date
from datetime import datetime
import sys
# =================================================================

os.system('cls' if os.name == 'nt' else 'clear')

# =================================================================
#QUAL PLANILHA VAI SER UTILIZADA?
planilha = "Pagamentos.xlsx"
#planilha = "Pagamentos - voluntariar abril 2026.xlsx"
# =================================================================

# =================================================================
#VARIÁVEIS IMPORTANTES
robo_deve_parar = False
linha = 2
numero_de_operacoes = 0
ainda_nao_foi_feito = '-'
loop1 = True
existe_planilha = False
processo = 0
despesa_certificada_teste = '-'
# =================================================================

# =================================================================
#PLANILHAS NO EXCEL:
while existe_planilha == False:
    #PLANILHA PRINCIPAL:
    try:
        book = openpyxl.load_workbook(planilha)
        pagina1 = book['Entrada']
        pagina2 = book['Despesas Certificadas']
        pagina3 = book['Notas de Liquidação']
        pagina4 = book['Preparações de Pagamento']
        pagina5 = book['Ordens Bancárias']
        pagina6 = book['Saída']
        existe_planilha = True
    except: 
        pyautogui.alert(text='\nDeu algum erro na planilha.', title='Erro', button='OK')
        sys.exit()
    #PLANILHA DE BACKUP:
    try:
        book_backup = openpyxl.load_workbook("Backup.xlsx")
        pagina1_backup = book_backup['Entrada']
        pagina2_backup = book_backup['Despesas Certificadas']
        pagina3_backup = book_backup['Notas de Liquidação']
        pagina4_backup = book_backup['Preparações de Pagamento']
        pagina5_backup = book_backup['Ordens Bancárias']
        pagina6_backup = book_backup['Saída']
        existe_planilha = True
    except:
        existe_planilha = False
        try:
            wb = openpyxl.Workbook()
            ws_principal = wb.active
            ws_principal.title = "Entrada"
            wb.save("Backup.xlsx")
            try:
                wb.create_sheet("Despesas Certificadas")
                wb.create_sheet("Notas de Liquidação")
                wb.create_sheet("Preparações de Pagamento")
                wb.create_sheet("Ordens Bancárias")
                wb.create_sheet('Saída')
                wb.save("Backup.xlsx")
                print("\nArquivo 'Backup.xlsx' criado com sucesso com várias planilhas.")
            except:
                print("\nErro na planilha de backup.")
        except:
            print("\nErro na planilha de backup.")

# =================================================================
print('=================================================================')
print('PROCEDIMENTO: CERTIFICAR E LIQUIDAR')
print('Versão 49 em 13/05/2026')
print('Por: LOUISE-SESDEC')
print('Github: https://github.com/louisegoncalves/Automatizacao-SIGEF')
print('=================================================================')
# =================================================================

#FUNÇÃO QUE SERÁ CHAMADA PELA TECLA DE PANICO
def parar_execucao():
    global robo_deve_parar
    print("\n!!! TECLA ESC ACIONADA! ENCERRANDO AUTOMACAO !!! \n")
    robo_deve_parar = True
    
#FUNÇÃO QUE ENCERRA O CODIGO E FECHA A PLANILHA COM SEGURANÇA
#A PLANILHA DEVE SEMPRE SER FECHADA ANTES DE ENCERRAR, POIS CORRE O RISCO DE CORROMPER
def verificar_panico_e_sair(workbook):
    global robo_deve_parar
    if robo_deve_parar:
        print("\nGarantindo que a planilha seja fechada...")
        if workbook:
            workbook.close()
        pyautogui.alert('\nTecla ESC acionada. Automação encerrada.')
        sys.exit()

#DEFINA SUA TECLA DE PÂNICO
tecla_de_panico = "Esc" 
keyboard.add_hotkey(tecla_de_panico, parar_execucao)
print(f"\nPressione a tecla '{tecla_de_panico}' a qualquer momento para abortar com segurança.")

#PORTA DO DEPURADOR DO GOOGLE CHROME
CHROME_DEBUG_URL = "http://localhost:9222"

if robo_deve_parar:
    verificar_panico_e_sair(book)

#DEFININDO NÚMERO DE OPERAÇÕES:
while processo != "None":
    processo = str(pagina1.cell(row=linha,column=3).value)
    numero_de_operacoes = numero_de_operacoes + 1
    linha = linha + 1
else: 
    linha = 2

numero_de_operacoes = numero_de_operacoes - 1

if numero_de_operacoes == 0:
    print("Nenhuma operação para executar, encerrando.")
else:
    
    #EXECUTANDO O PLAYWRIGHT DE FORMA SÍNCRONA
    with sync_playwright() as p:

            #CONECTAR AO NAVEGADOR JÁ ABERTO:
            browser_conectado = False
            while browser_conectado == False:
                print(f"\nTentando se conectar ao Chrome na porta de depuração: {CHROME_DEBUG_URL}")
                try:
                    browser = p.chromium.connect_over_cdp(CHROME_DEBUG_URL)
                    print("\nConexão estabelecida com sucesso!")
                    browser_conectado = True
                except: 
                    #AQUI ELE VAI PEDIR PARA ABRIR O SIGEF PELO DEPURADOR DO GOOGLE
                    pyautogui.confirm(text='Não identifiquei o SIGEF aberto. \n \nAperte OK quando o SIGEF estiver logado no depurador do Google Chrome', title='Depurador do Chrome' , buttons=['OK'])

            #OBTER A PÁGINA QUE ESTÁ ABERTA:
            janela = browser.contexts[0]
            guia = janela.pages[0]

            for p in janela.pages:
                if "SIGEF - Sistema Integrado de Planejamento e Gestão Fiscal" in p.title():
                    guia = p
                    break
            
            if guia is None:
                print("[ERRO]: Não encontrei nenhuma aba aberta com o SIGEF logado!")

            #VERIFICAR A PÁGINA ABERTA:
            print(f"\nAssumindo o controle da página com o título: '{guia.title()}'")
                
            #LOCALIZANDO O IFRAME:
            frame = guia.frame_locator('iframe[src="/SIGEF2026/SEG/#/SEGControleAcesso?p=1"]')
            
            #VERIFICAÇÃO DE PÂNICO:
            if robo_deve_parar:
                verificar_panico_e_sair(book)
            #SE QUISER DESATIVAR AQUELA JANELA DO COMEÇO PODE EXCLUIR ELA AQUI:
            pyautogui.alert(text='Procedimento: Certificar e liquidar. \n \nVersão 49 em 13/05/2026. \n \nPor: LOUISE-SESDEC', title='Início', button='OK')
# =================================================================

            os.system('cls' if os.name == 'nt' else 'clear')
            #INÍCIO                
            print("\n[INÍCIO DAS CERTIFICAÇÕES] Iniciando as certificações.")
            print("\nPlanilha lida. São " + str(numero_de_operacoes) + " operações para executar.")


            #PESQUISANDO FUNCIONALIDADE NO SIGEF
            pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
            pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
            pesquisar_funcionalidades_sistema.press_sequentially("Manter Despesa Certificada")
            funcionalidade_sistema = frame.get_by_title("Manter Despesa Certificada")
            
            #ABRINDO JANELA MANTER DESPESA CERTIFICADA:
            with guia.expect_popup() as popup_info:
                funcionalidade_sistema.click()
                manter_despesa_certificada = popup_info.value
            
            linha_documento = linha - 1

            #INÍCIO DO LOOP
            while numero_de_operacoes != linha_documento:

                #OBTENDO A UNIDADE GESTORA:
                #SE O CAMPO NÃO FOR PREENCHIDO, SERÁ ADOTADO POR PADRÃO A UNIDADE GESTORA 150001 (SESDEC)
                ug = str(pagina1.cell(row=linha, column=1).value)
                if ug == 'None':
                    ug = "150001"

                #OBTENDO A GESTÃO:
                #SE O CAMPO NÃO FOR PREENCHIDO, SERÁ ADOTADO POR PADRÃO A GESTÃO 00001 
                gestao = str(pagina1.cell(row=linha, column=2).value)
                if gestao == 'None':
                    gestao = "00001"

                #OBTENDO NÚMERO DO PROCESSO:
                processo = str(pagina1.cell(row=linha,column=3).value)
                if processo == 'None':
                    print('\n[ATENÇÃO] NÚMERO DO PROCESSO é inválido.')
                    break
                else:
                    if isinstance(processo,str):
                        processo = processo.replace('.','')
                        processo = processo.replace('-','')
                        processo = processo.replace('/','')
                        processo_sem_pontos = str(processo)

                        #AQUI SELECIONAMOS O NÚMERO DO MEIO DO PROCESSO:
                        processo_formatado = re.sub(r'(\d{4})(\d{6})(\d{4})(\d{2})', r'\1.\2/\3-\4', "{:016d}".format(int(processo_sem_pontos))) 
                        processo_cortado = processo_formatado.strip().split('/')[0]
                        processo_cortado = processo_cortado.strip().split('.')[1]
                
                #OBTENDO NOME COMPLETO DO FAVORECIDO:
                nome = str(pagina1.cell(row=linha, column=4).value)
                if nome == 'None':
                    print('\n[ATENÇÃO] NOME DO FAVORECIDO é inválido.')
                    break
                else:
                    primeiro_nome = nome.split()[0]
                
                #OBTENDO CPF DO FAVORECIDO:
                cpf = str(pagina1.cell(row=linha, column=5).value)
                if cpf == 'None':
                    print('\n[ATENÇÃO] CPF é inválido.')
                    break
                else:
                    #FORMATANDO O CAMPO DA PLANILHA "CPF" NO MOLDE 000.000.000-00:
                    if isinstance(cpf,str):
                        cpf = cpf.replace(' ','')
                        cpf = cpf.replace('.','')
                        cpf = cpf.replace('-','')
                        cpf_sem_ponto_virgula = str(cpf)
                        cpf_formatado = re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', "{:011d}".format(int(cpf_sem_ponto_virgula)))  
                        cpf = cpf_formatado
                    else: 
                        print('\n[ATENÇÃO] CPF é inválido.')
                        break
                
                #OBTENDO VALOR A SER PAGO AO FAVORECIDO:
                valor = str(pagina1.cell(row=linha, column=6).value)
                if valor == 'None':
                    print('\n[ATENÇÃO] VALOR é inválido.')
                    break
                else:
                    valor = valor.replace('R$','')
                    valor = valor.replace(' ','')
                    valor = valor.replace('.','')
                
                #OBTENDO BANCO, AGÊNCIA E CONTA:
                banco = str(pagina1.cell(row=linha, column=7).value)
                agencia = str(pagina1.cell(row=linha, column=8).value)
                conta = str(pagina1.cell(row=linha, column=9).value)

                #OBTENDO NOTA DE EMPENHO:
                empenho = str(pagina1.cell(row=linha, column=10).value)
                if empenho == 'None':
                    empenho = "Não definido"
                    exercicio = "2026"
                    print("[ATENÇÃO] Nota de empenho não foi definida.")
                    print("[ATENÇÃO] O robô prosseguirá somente até a etapa de certificação.")
                else:      
                    try:
                        exercicio = "2026"
                        empenho = int(empenho)
                    except:
                        empenho = str(empenho)
                    if isinstance(empenho,str):
                        empenho = empenho.upper()
                        exercicio = empenho.strip().split('NE')[0]
                        empenho = empenho.strip().split('NE')[1]
                        exercicio = int(exercicio)
                        empenho = int(empenho)
                        exercicio_NE = str(exercicio) + "NE"
                        nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                        nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
                    else: 
                        exercicio_NE = str(exercicio) + "NE"
                        nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                        nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)

                #OBTENDO DESPESA CERTIFICADA:
                despesa_certificada = str(pagina2.cell(row=linha, column=11).value)
                if despesa_certificada == 'None':
                    ja_foi_certificado = False
                else:
                    ja_foi_certificado = True
                
                #OBTENDO NOTA DE LIQUIDAÇÃO:
                liquidacao = str(pagina3.cell(row=linha, column=12).value)
                if liquidacao == '-':
                    ja_foi_liquidado = False
                else:
                    ja_foi_liquidado = True

                #OBTENDO PREPARAÇÃO DE PAGAMENTO:
                preparacao_pagamento = str(pagina4.cell(row=linha, column=13).value)

                #OBTENDO ORDEM BANCÁRIA:
                ordem_bancaria = str(pagina5.cell(row=linha, column=14).value)
                
                #OBTENDO A DATA DA ATIVIDADE EM QUE O FAVORECIDO FAZ JUS:
                data = str(pagina1.cell(row=linha, column=15).value)

                #OBTENDO A OPERAÇÃO/TIPO DE ATIVIDADE:
                operacao = str(pagina1.cell(row=linha, column=16).value)

                #OBTENDO A DATA QUE DEVERÁ SER EXECUTADO O PAGAMENTO:
                #SE NÃO FOR PREENCHIDO, ADOTARÁ POR PADRÃO A DATA DE HOJE.
                data_do_pagamento = str(pagina1.cell(row=linha, column=17).value)
                if data_do_pagamento == 'None':
                    data_atual = date.today() 
                    data_formatada = data_atual.strftime("%d/%m/%Y")
                    data_foi_formatada = True
                else:
                    if isinstance(data_do_pagamento,str):          
                        try:
                            data_formatada = datetime.strptime(data_do_pagamento, "%Y-%m-%d %H:%M:%S")                                    
                            data_formatada = data_formatada.strftime("%d/%m/%Y")
                            data_foi_formatada = True
                        except:
                            data_atual = date.today() 
                            data_formatada = data_atual.strftime("%d/%m/%Y")
                            data_foi_formatada = True

                #OBTENDO O NÚMERO DO DOCUMENTO:
                value_numero_cortado = str(pagina2.cell(row=linha, column=19).value)  
                if value_numero_cortado == 'None':
                    print("\nEstou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".") 
                    linha_documento = int(linha) - 1
                    value_numero_cortado = str(processo_cortado) + "-" + str(linha_documento)  
                else:
                    linha_documento = value_numero_cortado.strip().split('-')[1]
                    linha_documento = int(linha_documento)
                    ja_foi_certificado = True

                if despesa_certificada == 'None':
                    ja_foi_certificado = False
                    
                    if ja_foi_certificado == False:
                        texto_da_ce =  "Certificação de Despesa: Pagamento para o(a) servidor(a) " + str(cpf_formatado) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo_formatado) + "."

                        if robo_deve_parar:
                            verificar_panico_e_sair(book)
                            manter_despesa_certificada.close()

                        #INFORMAÇÕES PRELIMINARES
                        #HORA:
                        agora = datetime.now()

                        #PREENCHENDO INFORMAÇÕES NO SIGEF:
                        manter_despesa_certificada.wait_for_load_state('networkidle', timeout=10000)
                        campo_gestao = manter_despesa_certificada.locator("#txtCdGestao_SIGEFPesquisa")
                        campo_gestao.wait_for()
                        campo_gestao.fill(gestao)
                        tipo_documento = manter_despesa_certificada.locator("#cmbCdTipoDocumento")
                        tipo_documento.select_option(label="Outros")
                        numero_documento = manter_despesa_certificada.locator("#txtNuDocumento")
                        numero_documento.fill(value_numero_cortado)
                        favorecido = manter_despesa_certificada.locator("#txtNmCredor_lnkBtnPesquisa")
                        valor_documento = manter_despesa_certificada.locator("#txtVlDocumento")
                        data_emissao = manter_despesa_certificada.locator("#txtDtEmissao_SIGEFData")
                        data_aceite = manter_despesa_certificada.locator("#txtDtAceite_SIGEFData")
                        data_apresentacao = manter_despesa_certificada.locator("#txtDtApresentacao_SIGEFData")
                        competencia = manter_despesa_certificada.locator("#cboMesComp")
                        observacao = manter_despesa_certificada.locator("#txtDeObservacao")
                        atestado = manter_despesa_certificada.get_by_role("checkbox", name="Sou responsável pelo atesto")
                        data_emissao.click()
                        data_emissao.fill(data_formatada)
                        data_aceite.click()
                        data_aceite.fill(data_formatada)
                        data_apresentacao.click()
                        data_apresentacao.fill(data_formatada)

                        if data_foi_formatada == True:

                            #1. Extrai o mês da data
                            mes = data_formatada.strip().split('/')[1]
                            mes = mes.strip().split('/')[0]

                            #2. Cria o dicionário mapeando o número (chave) para o nome (valor)
                            meses = {
                                "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
                                "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
                                "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
                            }
                            
                            selecionar_competencia = meses.get(mes, 'Janeiro')
                            data_foi_formatada = False
                        
                        competencia.select_option(label=selecionar_competencia)
                        atestado.click()
                        
                        with manter_despesa_certificada.expect_popup() as popup_info:
                                favorecido.click()
                                selecionar_favorecido = popup_info.value

                                if robo_deve_parar:
                                    verificar_panico_e_sair(book)
                                    pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                                    manter_despesa_certificada.close()
                                selecionar_favorecido.wait_for_load_state('networkidle', timeout=10000)
                                botao_cpf = selecionar_favorecido.locator("#btnCPF")
                                botao_cpf.click()
                                preencher_cpf = selecionar_favorecido.get_by_role("textbox")
                                preencher_cpf.wait_for()
                                cpf_sem_ponto_virgula = str(cpf_sem_ponto_virgula)
                                preencher_cpf.fill(cpf_sem_ponto_virgula)
                                botao_confirmar = selecionar_favorecido.get_by_role("button", name="Confirmar a Consulta")
                                botao_confirmar.click()
                                selecionar_favorecido.wait_for_load_state('networkidle', timeout=10000)
                                localizar_funcao = selecionar_favorecido.get_by_text("* CPF")
                                localizar_funcao.wait_for()
                                try:
                                    codigo = selecionar_favorecido.get_by_role("cell", name=cpf_formatado, exact=True)
                                    codigo.wait_for()
                                    
                                    try:
                                        padrao_cpf = re.compile(r"\d{3}\.\d{3}\.\d{3}-\d{2}")
                                        primeira_celula_cpf = selecionar_favorecido.get_by_text(padrao_cpf).first
                                        primeira_celula_cpf.wait_for(timeout=10000)
                                        linha_correta = primeira_celula_cpf.locator("xpath=..")
                                        celula_nome_credor = linha_correta.locator("td").nth(1)
                                        nome_completo_na_tela = celula_nome_credor.inner_text()
                                        nome_completo_na_tela = nome_completo_na_tela.upper()
                                        primeiro_nome_na_tela = ""
                                        if nome_completo_na_tela and nome_completo_na_tela.strip():
                                            primeiro_nome_na_tela = nome_completo_na_tela.strip().split()[0]
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.upper()
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ç','C')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ã','A')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Á','A')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('À','A')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Í','I')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ô','O')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ô','O')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('É','E')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ú','U')
                                            primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ê','E')
                                            primeiro_nome = primeiro_nome.upper()
                                            primeiro_nome = primeiro_nome.replace('Ç','C')
                                            primeiro_nome = primeiro_nome.replace('Ã','A')
                                            primeiro_nome = primeiro_nome.replace('À','A')
                                            primeiro_nome = primeiro_nome.replace('Â','A')
                                            primeiro_nome = primeiro_nome.replace('Á','A')
                                            primeiro_nome = primeiro_nome.replace('Í','I')
                                            primeiro_nome = primeiro_nome.replace('Ô','O')
                                            primeiro_nome = primeiro_nome.replace('Õ','O')
                                            primeiro_nome = primeiro_nome.replace('É','E')      
                                            primeiro_nome = primeiro_nome.replace('Ú','U')    
                                            primeiro_nome = primeiro_nome.replace('Ê','E')           
                                        
                                            if primeiro_nome_na_tela.upper() == primeiro_nome.upper():
                                                print("[VALIDAÇÃO] Esperado " + primeiro_nome + ", encontrado " + primeiro_nome_na_tela + ".")
                                                codigo.click()
                                                print("[SUCESSO] Credor selecionado com sucesso.")

                                                validacao_bem_sucedida = True
                                                
                                            else:
                                                print("[ERRO DE VALIDAÇÃO] O nome não corresponde ao esperado!")
                                                validacao_bem_sucedida = False
                                                raise Exception('[ERRO DE VALIDAÇÃO] Esperado ' + primeiro_nome + " , encontrado " + primeiro_nome_na_tela + ".")
                                                
                                    except Exception as e:
                                            print(f"Ocorreu um erro durante a validação do credor: {e}")
                                
                                except:
                                    print("[ATENÇÃO] Não encontrei o CPF")
                                    todos_os_textos = codigo.all_inner_texts()
                                    numeros_pc = []
                                        
                                    if not numeros_pc:
                                        raise Exception("[ATENÇÃO] Nenhum número de CPF válido foi encontrado na lista de células.")
                        
                        if validacao_bem_sucedida == True:
                            manter_despesa_certificada.wait_for_load_state('networkidle', timeout=10000)
                            valor_documento.fill(valor)
                            observacao.fill(texto_da_ce)
                            botao_incluir = manter_despesa_certificada.get_by_role("button", name="Incluir o Registro")
                            if robo_deve_parar:
                                manter_despesa_certificada.close()
                                verificar_panico_e_sair(book)
                                manter_despesa_certificada.close()
                                    
                                pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                                sys.exit()
                            
                            cpf_final = manter_despesa_certificada.locator("#txtCdCredor").input_value()
                            credor_final = manter_despesa_certificada.locator("#txtNmCredor_SIGEFPesquisa").input_value()
                            credor_final = credor_final.upper()
                            valor_final = manter_despesa_certificada.locator("#txtVlDocumento").input_value()
                            botao_incluir.click()
                            manter_despesa_certificada.wait_for_load_state('networkidle', timeout=10000)
                            
                            try: 
                                erro_na_tela = manter_despesa_certificada.get_by_role("cell", name="Número Documento já cadastrado(a).", exact=True)
                                if erro_na_tela.is_visible():
                                    documento_ja_cadastrado = True
                                else:
                                    documento_ja_cadastrado = False
                                if documento_ja_cadastrado:
                                    print("[AVISO] O documento já foi cadastrado anteriormente.")
                                    print("O robô vai pular este item ou tomar uma ação alternativa.")
                                else:
                                    print("[SUCESSO] Nenhuma mensagem de erro encontrada.")
                                    documento_ja_cadastrado = False
                            except Exception as e:
                                documento_ja_cadastrado = False
                                print(f"Ocorreu um erro durante a verificação do documento: {e}")
                            
                            if documento_ja_cadastrado == True:
                                try:
                                    despesa_certificada = "pesquisar no sigef"
                                    pagina2_backup.append([ug,gestao,processo_formatado,credor_final,cpf_final,valor_final,banco,agencia,conta,nota_de_empenho,despesa_certificada,ainda_nao_foi_feito,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_formatada,agora,value_numero_cortado])
                                    book_backup.save("Backup.xlsx") 
                                    
                                    dados = [ug,gestao,processo_formatado,credor_final,cpf_final,valor_final,banco,agencia,conta,nota_de_empenho,despesa_certificada,ainda_nao_foi_feito,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_formatada,agora,value_numero_cortado]

                                    for numero_coluna, valor in enumerate(dados, start=1):
                                        pagina2.cell(row=linha, column=numero_coluna, value=valor)

                                    book.save(planilha)
                                    
                                except:
                                    book_backup.save("Backup.xlsx") 
                                    print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                    book_backup.close()
                                    sys.exit()
                            else:
                                numero_despesa_certificada = manter_despesa_certificada.locator("#txtNuSeq")
                                numero_despesa_certificada.wait_for(timeout=10000)
                                numero_despesa_certificada.dblclick()
                                numero_despesa_certificada.press('Control+KeyC')
                                despesa_certificada =  numero_despesa_certificada = pyperclip.paste()
                                despesa_certificada = "2026CE" + str(despesa_certificada)
                                print(f"[SUCESSO] Despesa Certificada encontrada e copiada: '{despesa_certificada}'")

                                if despesa_certificada_teste == despesa_certificada:
                                    print("[REPETIDO] Refazendo Despesa Certificada!")
                                    despesa_certificada = 'None'
                                else:
                                    if despesa_certificada_teste == despesa_certificada:
                                        print("[ERRO] Refazendo Despesa Certificada!")
                                        book.save(planilha)
                                        despesa_certificada = 'None'
                                    else:
                                        if despesa_certificada == "pesquisar no sigef":
                                            time.sleep(0)
                                        else:
                                            try:
                                                pagina2_backup.append([ug,gestao,processo_formatado,credor_final,cpf_final,valor_final,banco, agencia,conta,nota_de_empenho,despesa_certificada,ainda_nao_foi_feito,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_formatada,agora,value_numero_cortado])
                                                book_backup.save("Backup.xlsx")
                        
                                                dados = [ug,gestao,processo_formatado,credor_final,cpf_final,valor_final,banco,agencia,conta,nota_de_empenho,despesa_certificada,ainda_nao_foi_feito,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_formatada,agora,value_numero_cortado]

                                                for numero_coluna, valor in enumerate(dados, start=1):
                                                    pagina2.cell(row=linha, column=numero_coluna, value=valor)

                                                book.save(planilha)
                                                
                                            except:
                                                book_backup.save("Backup.xlsx")
                                                print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                                book_backup.close()
                                                sys.exit()
                        else:
                            botao_limpar = manter_despesa_certificada.get_by_role("link", name="Limpar a Tela")
                            botao_limpar.click()
                            cpf = "CPF não condiz com o servidor"
                            despesa_certificada = "Não foi feita."
                            ug = gestao = processo = valor = banco = agencia = conta = empenho = liquidacao = operacao = data = data_do_pagamento = agora = data_formatada = value_numero_cortado = '-'
                            dados = [ug,gestao,processo,nome,cpf,valor,banco,agencia,conta,empenho,despesa_certificada,ainda_nao_foi_feito,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_formatada, agora,value_numero_cortado]

                            for numero_coluna, valor in enumerate(dados, start=1):
                                pagina2.cell(row=linha, column=numero_coluna, value=valor)

                            book.save(planilha)

                        botao_limpar = manter_despesa_certificada.get_by_role("link", name="Limpar a Tela")
                        botao_limpar.click()

                else:
                    ja_foi_certificado = True

                if despesa_certificada != 'None':
                    despesa_certificada_teste = despesa_certificada 
                    linha = linha + 1
                    ug = gestao = processo = nome = cpf = valor = banco = agencia = conta = empenho = despesa_certificada = liquidacao = operacao = data = data_do_pagamento = agora = value_numero_cortado = 'None'
                    documento_ja_cadastrado = False
    
            print("\n[FIM DAS CERTIFICAÇÕES] Nenhuma despesa para certificar.")
            ja_foi_certificado = True
            manter_despesa_certificada.close()

##################################################################################################################################
#              A PARTIR DAQUI COMEÇA A LIQUIDAR
##################################################################################################################################
            print("\n=====================================================")
            linha = 2
            linha_documento = linha
        
            try:
                book = openpyxl.load_workbook(planilha)
                pagina1 = book['Entrada']
                pagina2 = book['Despesas Certificadas']
                pagina3 = book['Notas de Liquidação']
                pagina4 = book['Preparações de Pagamento']
                pagina5 = book['Ordens Bancárias']
                pagina6 = book['Saída']
            except: 
                pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
                sys.exit()

            pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
            pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
            pesquisar_funcionalidades_sistema.press_sequentially("Liquidar Despesa Certificada")
            funcionalidade_sistema = frame.get_by_title("Liquidar Despesa Certificada")

            print("\n[INÍCIO DAS LIQUIDAÇÕES] Iniciando as liquidações.")
                            
            #INÍCIO DO LOOP
            while numero_de_operacoes != linha_documento:
                
                linha_documento = linha - 1

                #OBTENDO CPF DO FAVORECIDO:
                cpf = str(pagina2.cell(row=linha, column=5).value)
                if cpf == 'None':
                    print('\n[ATENÇÃO] CPF é inválido.')
                    break
                else:
                    with guia.expect_popup() as popup_info:

                        funcionalidade_sistema.click()
                        liquidar_despesa_certificada = popup_info.value

                        while ja_foi_certificado == True:

                            #OBTENDO A UNIDADE GESTORA:
                            ug = str(pagina2.cell(row=linha, column=1).value)
                            
                            #OBTENDO A GESTÃO:
                            gestao = str(pagina2.cell(row=linha, column=2).value)

                            #OBTENDO NÚMERO DO PROCESSO:
                            processo = str(pagina2.cell(row=linha, column=3).value)
                            
                            #OBTENDO NOME COMPLETO DO FAVORECIDO:
                            nome = str(pagina2.cell(row=linha, column=4).value)

                            #OBTENDO CPF DO FAVORECIDO:
                            cpf = str(pagina2.cell(row=linha, column=5).value)

                            if nome == 'None':
                                print('\n[ATENÇÃO] NOME DO FAVORECIDO é inválido.')
                                break
                            else:
                                primeiro_nome = nome.split()[0]
                            
                            #OBTENDO VALOR A SER PAGO AO FAVORECIDO:
                            valor = str(pagina2.cell(row=linha, column=6).value)
                            
                            #OBTENDO BANCO, AGÊNCIA E CONTA:
                            banco = str(pagina1.cell(row=linha, column=7).value)
                            agencia = str(pagina1.cell(row=linha, column=8).value)
                            conta = str(pagina1.cell(row=linha, column=9).value)

                            #OBTENDO NOTA DE EMPENHO:
                            empenho = str(pagina2.cell(row=linha, column=10).value)
                            if empenho == "Não definido":
                                print("[ATENÇÃO] Nota de empenho não foi definida.")
                                print("[ATENÇÃO] O robô não prosseguirá enquanto não for definida a nota de empenho na planilha.")
                                break
                            else:
                                nota_de_empenho = empenho
                                exercicio = empenho.strip().split('NE')[0]
                                empenho = empenho.strip().split('NE')[1]
                            
                            #OBTENDO DESPESA CERTIFICADA:
                            despesa_certificada = str(pagina2.cell(row=linha, column=11).value)
                            if despesa_certificada == 'None':
                                ja_foi_certificado = False
                            else:
                                ja_foi_certificado = True
                        
                            #OBTENDO NOTA DE LIQUIDAÇÃO:
                            liquidacao = str(pagina3.cell(row=linha, column=12).value)
                            if liquidacao == 'None':
                                ja_foi_liquidado = False
                            else:
                                ja_foi_liquidado = True

                            #OBTENDO PREPARAÇÃO DE PAGAMENTO:
                            preparacao_pagamento = str(pagina4.cell(row=linha, column=13).value)

                            #OBTENDO ORDEM BANCÁRIA:
                            ordem_bancaria = str(pagina5.cell(row=linha, column=14).value)
                            
                            #OBTENDO A DATA DA ATIVIDADE EM QUE O FAVORECIDO FAZ JUS:
                            data = str(pagina1.cell(row=linha, column=15).value)

                            #OBTENDO A OPERAÇÃO/TIPO DE ATIVIDADE:
                            operacao = str(pagina1.cell(row=linha, column=16).value)

                            #OBTENDO A DATA QUE DEVERÁ SER EXECUTADO O PAGAMENTO:
                            #SE NÃO FOR PREENCHIDO, ADOTARÁ POR PADRÃO A DATA DE HOJE.
                            data_do_pagamento = str(pagina2.cell(row=linha, column=17).value)

                            #OBTENDO O NÚMERO DO DOCUMENTO:
                            value_numero_cortado = str(pagina2.cell(row=linha, column=19).value) 

                            while ja_foi_liquidado == False:
                                texto_da_nl = "Liquidação de Despesa: Pagamento para o(a) servidor(a) " + str(cpf) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo) + "."
                                print("\nEstou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ", despesa certificada " + str(despesa_certificada) + ".")

                                if robo_deve_parar:
                                    liquidar_despesa_certificada.close()
                                    verificar_panico_e_sair(book)
                                    
        
                                liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=10000)
                                campo_unidade_gestora = liquidar_despesa_certificada.locator("#txtCdUnidadeGestora")
                                campo_unidade_gestora.wait_for()
                                campo_unidade_gestora.fill(ug)
                                campo_gestao = liquidar_despesa_certificada.locator("#txtCdGestao_SIGEFPesquisa")

                                #INFORMAÇÕES PRELIMINARES
                                #HORA EM QUE ESTÁ SENDO EXECUTADA A OPERAÇÃO:
                                agora = datetime.now()
                                
                                if robo_deve_parar:
                                    liquidar_despesa_certificada.close()
                                    if book:
                                        verificar_panico_e_sair(book)

                                campo_gestao.fill(gestao)
                                campo_despesa_certificada = liquidar_despesa_certificada.locator("#txtDespesaCertificadaNumero_SIGEFPesquisa")
                                ce = despesa_certificada.replace("2026CE","")
                                campo_despesa_certificada.fill(ce)
                                botao_pesquisar = liquidar_despesa_certificada.get_by_role("button", name="Confirmar a Pesquisa")
                                botao_pesquisar.click()
                                    
                                if robo_deve_parar:
                                    liquidar_despesa_certificada.close()
                                    if book:
                                        verificar_panico_e_sair(book)
                                    
                                data_vencimento = liquidar_despesa_certificada.locator("#txtDataVencimento_SIGEFData")
                                data_vencimento.fill(data_formatada)
                                adicionar = liquidar_despesa_certificada.get_by_role("button", name="Adicionar Documento")
                                adicionar.click()
                                liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=10000)
                                data_referencia = liquidar_despesa_certificada.locator("#txtDtReferenciaId_SIGEFData")
                                data_referencia.wait_for()
                                data_referencia.fill(data_formatada)
                                ponto_interrogacao = liquidar_despesa_certificada.locator("#txtNotaEmpenhoNumeroId_lnkBtnPesquisa")

                                if robo_deve_parar:
                                    liquidar_despesa_certificada.close()
                                    if book:
                                        verificar_panico_e_sair(book)
                                        
                                with liquidar_despesa_certificada.expect_popup() as popup_info:
                                    ponto_interrogacao.click()
                                    if robo_deve_parar:
                                        liquidar_despesa_certificada.close()
                                        if book:
                                            verificar_panico_e_sair(book)
                                    selecionar_empenho = popup_info.value
                                    selecionar_empenho.wait_for_load_state('networkidle', timeout=30000)
                                    preencher_empenho_ano = selecionar_empenho.locator("#txtNotaEmpenhoAno")
                                    preencher_empenho = selecionar_empenho.locator("#txtNotaEmpenhoNumero")
                                    preencher_empenho.fill(empenho)
                                    preencher_empenho_ano.fill(exercicio)

                                    botao_confirmar = selecionar_empenho.get_by_role("button", name="Confirmar a Consulta")
                                    botao_confirmar.click()
                                    selecionar_empenho.wait_for_load_state('networkidle', timeout=30000)
                                    nota_empenho = selecionar_empenho.get_by_role("cell", name=nota_de_empenho, exact=True)
                                    nota_empenho.wait_for()
                                    nota_empenho.click()
                                    time.sleep(0.5)
                                    liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                                    valor_bruto = liquidar_despesa_certificada.locator("#txtValorBrutoId")
                                    valor_bruto.wait_for()
                                    valor_bruto.fill(valor)
                                    botao_retencoes = liquidar_despesa_certificada.get_by_role("button", name="Sugerir Retenções")
                                    botao_retencoes.click()
                                    nao_existem_retencoes = liquidar_despesa_certificada.get_by_text("Não existem sugestões para")
                                    nao_existem_retencoes.wait_for()
                                    valor_liquido = liquidar_despesa_certificada.locator("#txtValorLiquidoId")
                                    valor_liquido.wait_for()
                                    valor_liquido = liquidar_despesa_certificada.locator("#txtValorLiquidoId").input_value()
                                    valor_que_vai_pra_planilha = valor_liquido

                                    valor_liquido = valor_liquido.replace(".","")
                                    valor_liquido = valor_liquido.replace(",","")
                                    valor = valor.replace(",","")
                                    valor = valor.replace(".","")
                                
                                    if valor_liquido == valor:

                                        liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                                            
                                        if robo_deve_parar:
                                            liquidar_despesa_certificada.close()
                                            if book:
                                                verificar_panico_e_sair(book)

                                        historico = liquidar_despesa_certificada.locator("#txtHistorico")
                                        historico.wait_for()
                                        historico.press('Control+KeyA')
                                        historico.press('Delete')
                                        historico.fill(texto_da_nl)
                                        botao_confirmacao = liquidar_despesa_certificada.locator("#menun4").get_by_role("link")
                                        
                                        if robo_deve_parar:
                                            liquidar_despesa_certificada.close()
                                            if book:
                                                verificar_panico_e_sair(book)
                                            
                                        botao_confirmacao.click()
                                        botao_confirmar = liquidar_despesa_certificada.get_by_role("button", name="Confirmar a Operação")
                                        botao_limpar = liquidar_despesa_certificada.get_by_role("link", name="Limpar a Tela")
                                            
                                        if robo_deve_parar:
                                            liquidar_despesa_certificada.close()
                                            if book:
                                                verificar_panico_e_sair(book)
                                
                                        try:
                                            botao_confirmar.click()
                                            try:
                                                erro = liquidar_despesa_certificada.get_by_text("Não é permitido liquidar da")
                                                erro_esta_visivel = erro.is_visible()
                                                if erro_esta_visivel:
                                                    print("[ERRO DE VALIDAÇÃO] Mensagem de erro detectada: 'Não é permitido liquidar da'.")
                                                    botao_voltar = liquidar_despesa_certificada.get_by_role("button", name="Retornar à Tela Anterior")
                                                    botao_voltar.click()
                                                    documento_ja_liquidado_mas_nao_salvo = True
                                                else:
                                                    documento_ja_liquidado_mas_nao_salvo = False
                                            except:
                                                    time.sleep(0)
                                        except Exception as e:
                                            print(f"[ERRO DE VALIDAÇÃO] Ocorreu um erro inesperado durante a liquidação: {e}")

                                        try:
                                            liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=10000)
                                            padrao_da_nl = re.compile(r"^2026NL\d{6}$")
                                            celula_nl = liquidar_despesa_certificada.get_by_role("cell", name=padrao_da_nl)
                                            celula_nl.first.wait_for(state="visible", timeout=10000)

                                            if celula_nl.count() > 0:
                                                    primeira_nl = celula_nl.first
                                                    liquidacao = primeira_nl.inner_text()
                                                    print(f"[SUCESSO] Nota de Liquidação gerada: '{liquidacao}'")
                                                    ja_foi_liquidado = True
                                            else:
                                                    liquidacao = "ERRO"
                                                    print("[AVISO] Nenhuma Nota de Liquidação foi encontrada na página.")

                                        except Exception as e:
                                                print(f"Ocorreu um erro ao tentar localizar a NL: {e}")
                                                liquidacao = "ERRO"

                                        if liquidacao == "ERRO":
                                    
                                            try:
                                                    pagina3_backup.append([ug,gestao,processo,nome,cpf,valor_que_vai_pra_planilha,banco, agencia,conta,nota_de_empenho,despesa_certificada,liquidacao,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado])
                                                    book_backup.save("Backup.xlsx")
                                                    
                                                    dados = [ug,gestao,processo,nome,cpf,valor_que_vai_pra_planilha,banco,agencia,conta,nota_de_empenho,despesa_certificada,liquidacao,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado]

                                                    for numero_coluna, valor in enumerate(dados, start=1):
                                                        pagina3.cell(row=linha, column=numero_coluna, value=valor)

                                                    book.save(planilha)
                                                    
                                            except:           
                                                    
                                                    book_backup.save("Backup.xlsx")
                                                    print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                                    book_backup.close()
                                                    sys.exit()
                                    else:
                                        print('[ATENÇÃO] Valor da planilha está diferente do valor que está no SIGEF!') 
                                        break
                                    
                                    try:
                                        botao_limpar.click()
                                        
                                        pagina3_backup.append([ug,gestao,processo,nome,cpf,valor_que_vai_pra_planilha,banco,agencia,conta,nota_de_empenho,despesa_certificada,liquidacao,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado])

                                        book_backup.save("Backup.xlsx")
                                        
                                        dados = [ug,gestao,processo,nome,cpf,valor_que_vai_pra_planilha,banco,agencia,conta,nota_de_empenho,despesa_certificada,liquidacao,ainda_nao_foi_feito,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado]

                                        for numero_coluna, valor in enumerate(dados, start=1):
                                            pagina3.cell(row=linha, column=numero_coluna, value=valor)

                                        book.save(planilha)

                                    except:                              
                                        book_backup.save("Backup.xlsx")
                                        print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                        book_backup.close()
                                        sys.exit()
                            else: 
                                linha = linha + 1
                                ug = gestao = processo = nome = cpf = valor = banco = agencia = conta = empenho = despesa_certificada = liquidacao = operacao = data = data_do_pagamento = agora = value_numero_cortado = 'None'

liquidar_despesa_certificada.close()
print("\n[FIM DAS LIQUIDAÇÕES] Nenhuma despesa para liquidar.")
if book:
    book.close()
print("\nScript finalizado. A janela de depuração permanece aberta.")
keyboard.remove_hotkey(tecla_de_panico) 
pyautogui.alert(text='Encerrei por aqui.', title='Fim', button='OK')
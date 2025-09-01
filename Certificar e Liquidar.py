#OLÁ!
#PROCEDIMENTO: CERTIFICAR E LIQUIDAR;
#POR: LOUISE-SESDEC;
#ALTERAÇÕES NO CÓDIGO PODEM SER ACESSADAS NO MEU GITHUB: <https://github.com/louisegoncalves/Automatizacao-SIGEF>.

#INSTRUÇÕES
#ATENÇÃO: É OBRIGATÓRIO ABRIR O DEPURADOR DO GOOGLE CHROME PARA EXECUTAR ESSE CÓDIGO
#EXECUTE NO CMD: "C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebugProfile"
#E LOGUE NO SIGEF

#BIBLIOTECAS UTILIZADAS:
from playwright.sync_api import sync_playwright, Page, TimeoutError
import pyautogui
import openpyxl
import keyboard
import time
import pyperclip
import re
from datetime import date
from datetime import datetime
import sys

#VARIÁVEIS IMPORTANTES
robo_deve_parar = False
coluna = 1
linha = 2
loop = True
despesa_certificada_teste = 'None'

#PLANILHA NO EXCEL:
try:
    book = openpyxl.load_workbook('Pagamentos.xlsx')
    pagina = book['Entrada']
    pagina1 = book['Despesas Certificadas']
    pagina2 = book['Notas de Liquidação']
    pagina3 = book['Preparações de Pagamento']
    pagina4 = book['Ordens Bancárias']
    pagina5 = book['Saída']
except: 
    pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
    sys.exit()

#SE QUISER DESATIVAR AQUELA JANELA DO COMEÇO PODE EXCLUIR ELA AQUI:
pyautogui.alert(text='Procedimento: Certificar e liquidar.', title='Início', button='OK')

#FUNÇÃO QUE SERÁ CHAMADA PELA TECLA DE PANICO
def parar_execucao():
    global robo_deve_parar
    print("\n!!! TECLA ESC ACIONADA! ENCERRANDO AUTOMACAO !!!")
    robo_deve_parar = True

#FUNÇÃO QUE ENCERRA O CODIGO E FECHA A PLANILHA COM SEGURANÇA
#A PLANILHA DEVE SEMPRE SER FECHADA ANTES DE ENCERRAR, POIS CORRE O RISCO DE CORROMPER
def verificar_panico_e_sair(workbook):
    global robo_deve_parar
    if robo_deve_parar:
        print("Garantindo que a planilha seja fechada...")
        if workbook:
            workbook.close()
        pyautogui.alert('Tecla ESC acionada. Automação encerrada.')
        sys.exit()

#DEFINA SUA TECLA DE PÂNICO
tecla_de_panico = "Esc" 
keyboard.add_hotkey(tecla_de_panico, parar_execucao)
print(f"Robô iniciado. Pressione a tecla '{tecla_de_panico}' a qualquer momento para abortar com seguranca.")

#AQUI ELE VAI PEDIR PARA ABRIR O SIGEF PELO DEPURADOR DO GOOGLE
pyautogui.confirm(text='Aperte OK quando o SIGEF estiver logado no depurador do Google Chrome', title='Depurador do Chrome' , buttons=['OK'])

#PORTA DO DEPURADOR DO GOOGLE CHROME
CHROME_DEBUG_URL = "http://localhost:9222"

if robo_deve_parar:
    verificar_panico_e_sair(book)

#EXECUTANDO O PLAYWRIGHT DE FORMA SÍNCRONA
with sync_playwright() as p:
        if robo_deve_parar:
            verificar_panico_e_sair(book)

        #CONECTAR AO NAVEGADOR JÁ ABERTO:
        print(f"Tentando se conectar ao Chrome na porta de depuração: {CHROME_DEBUG_URL}")
        browser = p.chromium.connect_over_cdp(CHROME_DEBUG_URL)
        print("Conexão estabelecida com sucesso!")

        #OBTER A PÁGINA QUE ESTÁ ABERTA:
        #Quando conectamos, precisamos pegar o contexto e a página existentes;
        #Geralmente, a primeira página do primeiro contexto é a que queremos.
        janela = browser.contexts[0]
        guia = janela.pages[0]

        #VERIFICAR A PÁGINA ABERTA:
        print(f"Assumindo o controle da página com o título: '{guia.title()}'")
            
        #LOCALIZANDO O IFRAME:
        frame = guia.frame_locator('iframe[src="/SIGEF2025/SEG/#/SEGControleAcesso?p=1"]')
        
        #VERIFICAÇÃO DE PÂNICO:
        if robo_deve_parar:
            verificar_panico_e_sair(book)
        
        #INÍCIO                
        print("Iniciando!")

        #PESQUISANDO FUNCIONALIDADE NO SIGEF
        pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
        pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
        pesquisar_funcionalidades_sistema.press_sequentially("Manter Despesa Certificada")
        funcionalidade_sistema = frame.get_by_title("Manter Despesa Certificada")
        
        #ABRINDO JANELA MANTER DESPESA CERTIFICADA:
        with guia.expect_popup() as popup_info:
            funcionalidade_sistema.click()
            manter_despesa_certificada = popup_info.value
        
        #INÍCIO DO LOOP
        while loop == True:
            
            # A COLUNA SEMPRE DEVE REINICIAR NO INÍCIO DO LOOP
            coluna = 1
            
            #LENDO A UG
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ug = cell.value
                    ug= str(ug)

            #LENDO A GESTÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    gestao = cell.value
                    gestao = str(gestao)

            #LENDO O Nº DO PROCESSO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    processo = cell.value
                    processo = str(processo)
            
            #LENDO O NOME DO SERVIDOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    nome = cell.value
                    nome = str(nome)
                    primeiro_nome = nome.split()[0]

            #LENDO O CPF
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    cpf = cell.value
                    cpf = str(cpf)
                    
            #LENDO O VALOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):                
                for cell in row:
                    valor = cell.value
                    valor = str(valor)
                    valor = valor.replace('.',',')
            
            #LENDO O BANCO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    banco = cell.value
                    banco = str(banco)

            #LENDO A AGENCIA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    agencia = cell.value
                    agencia = str(agencia)

            #LENDO A CONTA CORRENTE
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    conta = cell.value
                    conta = str(conta)

            #LENDO A NOTA DE EMPENHO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    empenho = cell.value
                    empenho = str(empenho)

            #LENDO A DESPESA CERTIFICADA
            coluna = coluna + 1
            for row in pagina1.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    despesa_certificada = cell.value
                    despesa_certificada = str(despesa_certificada)

            #LENDO A NOTA DE LIQUIDAÇÃO
            coluna = coluna + 1
            for row in pagina2.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    liquidacao = cell.value
                    liquidacao = str(liquidacao)
            
            #LENDO A PREPARAÇÃO DE PAGAMENTO
            coluna = coluna + 1
            for row in pagina3.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    preparacao_pagamento = cell.value
                    preparacao_pagamento = str(preparacao_pagamento)

            #LENDO A ORDEM BANCÁRIA
            coluna = coluna + 1
            for row in pagina4.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ordem_bancaria = cell.value
                    ordem_bancaria = str(ordem_bancaria)

            #LENDO A DATA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    data = cell.value
                    data = str(data)

            #LENDO A OPERAÇÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    operacao = cell.value
                    operacao = str(operacao)
            
            #FORMATANDO O CAMPO DA PLANILHA "CPF" NO MOLDE 000.000.000-00
            #O RESULTADO SERÁ DUAS VARIÁVEIS: cpf_sem_ponto_virgula e cpf_formatado
            if cpf != "None":
                try:
                    if isinstance(cpf,str):
                        cpf = cpf.replace('.','')
                        cpf = cpf.replace('-','')
                        cpf_sem_ponto_virgula = str(cpf)
                    else: 
                        print('CPF é inválido.')
    
                    cpf_formatado = re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', "{:011d}".format(int(cpf_sem_ponto_virgula)))                        
                    cpf = cpf_formatado
                except:
                    time.sleep(0)    
                   
            if nome != "None":
                print("Estou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".")
            else:
                time.sleep(0)

            if nome == "None":
                if valor == "None":
                    loop = False
                    break
            
            if isinstance(processo,str):
                processo = processo.replace('.','')
                processo = processo.replace('-','')
                processo = processo.replace('/','')
                processo_sem_pontos = str(processo)

            else: 
                print('Processo é inválido.')
            
            processo_formatado = re.sub(r'(\d{4})(\d{6})(\d{4})(\d{2})', r'\1.\2/\3-\4', "{:016d}".format(int(processo_sem_pontos))) 
            #AQUI SELECIONAMOS O NÚMERO DO MEIO DO PROCESSO:
            processo_cortado = processo_formatado.strip().split('/')[0]
            processo_cortado = processo_cortado.strip().split('.')[1]
            value_numero_cortado = str(processo_cortado) + "-" + str(linha)
            
            
            try:
                exercicio = int(exercicio)
            except:
                try:
                    empenho = int(empenho)
                except:
                    exercicio = '2025'
                    empenho = str(empenho)

            if isinstance(empenho,str):
                empenho = empenho.upper()
                exercicio = empenho.strip().split('NE')[0]
                empenho = empenho.strip().split('NE')[1]
                exercicio = int(exercicio)
                empenho = int(empenho)
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
            else: 
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)

            if despesa_certificada == "None":
                ja_foi_certificado = False
                if ja_foi_certificado == False:
                    texto_da_ce =  "Certificação de Despesa: Pagamento para o(a) servidor(a) " + str(cpf_formatado) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo_formatado) + "."
                    texto_da_ce = texto_da_ce.replace('ç','c')
                    texto_da_ce = texto_da_ce.replace('ã','a')
                    texto_da_ce = texto_da_ce.replace('á','a')
                    texto_da_ce = texto_da_ce.replace('à','a')
                    texto_da_ce = texto_da_ce.replace('í','i')
                    texto_da_ce = texto_da_ce.replace('ô','o')
                    texto_da_ce = texto_da_ce.replace('õ','o')
                    texto_da_ce = texto_da_ce.replace('é','e')

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    #INFORMAÇÕES PRELIMINARES
                    #DATA DE HOJE:
                    data_atual = date.today() 
                    #FORMATAR A DATA:
                    data_de_baixa = data_atual.strftime("%d/%m/%Y")
                    #HORA:
                    agora = datetime.now()
                    #FORMATAR A HORA:
                    hora = agora.strftime("%H:%M:%S")
                    manter_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    time.sleep(0.1)
                    campo_gestao = manter_despesa_certificada.locator("#txtCdGestao_SIGEFPesquisa")
                    campo_gestao.wait_for()
                    campo_gestao.press_sequentially(gestao)
                    tipo_documento = manter_despesa_certificada.locator("#cmbCdTipoDocumento")
                    tipo_documento.select_option(label="Outros")
                    numero_documento = manter_despesa_certificada.locator("#txtNuDocumento")
                    numero_documento.press_sequentially(value_numero_cortado)
                    favorecido = manter_despesa_certificada.locator("#txtNmCredor_lnkBtnPesquisa")
                    valor_documento = manter_despesa_certificada.locator("#txtVlDocumento")
                    data_emissao = manter_despesa_certificada.locator("#txtDtEmissao_SIGEFData")
                    data_aceite = manter_despesa_certificada.locator("#txtDtAceite_SIGEFData")
                    data_apresentacao = manter_despesa_certificada.locator("#txtDtApresentacao_SIGEFData")
                    competencia = manter_despesa_certificada.locator("#cboMesComp")
                    observacao = manter_despesa_certificada.locator("#txtDeObservacao")
                    atestado = manter_despesa_certificada.get_by_role("checkbox", name="Sou responsável pelo atesto")
                    time.sleep(0.2)
                    data_emissao.click()
                    data_emissao.press_sequentially(data_de_baixa)
                    time.sleep(0.2)
                    data_aceite.click()
                    data_aceite.press_sequentially(data_de_baixa)
                    time.sleep(0.2)
                    data_apresentacao.click()
                    data_apresentacao.press_sequentially(data_de_baixa)
                    time.sleep(0.2)
                    competencia.select_option(label="Agosto")
                    time.sleep(0.2)
                    atestado.click()
                    
                    with manter_despesa_certificada.expect_popup() as popup_info:
                            favorecido.click()
                            selecionar_favorecido = popup_info.value

                            if robo_deve_parar:
                                pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                                
                            selecionar_favorecido.wait_for_load_state('networkidle', timeout=30000)
                            time.sleep(0.5)
                            botao_cpf = selecionar_favorecido.locator("#btnCPF")
                            botao_cpf.click()
                            preencher_cpf = selecionar_favorecido.get_by_role("textbox")
                            preencher_cpf.wait_for()
                            cpf_sem_ponto_virgula = str(cpf_sem_ponto_virgula)
                            preencher_cpf.press_sequentially(cpf_sem_ponto_virgula)
                            time.sleep(0.5)
                            botao_confirmar = selecionar_favorecido.get_by_role("button", name="Confirmar a Consulta")
                            botao_confirmar.click()
                            selecionar_favorecido.wait_for_load_state('networkidle', timeout=30000)
                            localizar_funcao = selecionar_favorecido.get_by_text("* CPF")
                            localizar_funcao.wait_for()
                            try:
                                codigo = selecionar_favorecido.get_by_role("cell", name=cpf_formatado, exact=True)
                                codigo.wait_for()
                                
                                try:
                                    padrao_cpf = re.compile(r"\d{3}\.\d{3}\.\d{3}-\d{2}")
    
                                    #AQUI ENCONTRAMOS A PRIMEIRA CÉLULA QUE CORRESPONDE AO PADRÃO DO CPF:
                                    primeira_celula_cpf = selecionar_favorecido.get_by_text(padrao_cpf).first
                                        
                                    #É crucial esperar que esta âncora apareça
                                    primeira_celula_cpf.wait_for(timeout=10000)
                                    print("Célula âncora encontrada com o texto: " + primeira_celula_cpf.inner_text())

                                    # --- PASSO 2: A PARTIR DA ÂNCORA, NAVEGAR ATÉ A LINHA PAI ---
                                    # O XPath '..' significa "vá para o elemento pai".
                                    # O pai de uma célula (<td>) é a sua linha (<tr>).
                                    linha_correta = primeira_celula_cpf.locator("xpath=..")
                                        
                                    # --- PASSO 3: DA LINHA, NAVEGAR ATÉ A CÉLULA DO NOME ---
                                    # Agora que temos a linha correta, pegamos a segunda célula (índice 1)
                                    celula_nome_credor = linha_correta.locator("td").nth(1)
                                        
                                    nome_completo_na_tela = celula_nome_credor.inner_text()
                                    nome_completo_na_tela = nome_completo_na_tela.upper()
                                        
                                    # --- PASSO 4: VALIDAR E AGIR ---
                                    primeiro_nome_na_tela = ""
                                    if nome_completo_na_tela and nome_completo_na_tela.strip():
                                        primeiro_nome_na_tela = nome_completo_na_tela.strip().split()[0]
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ç','C')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ã','A')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Á','A')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('À','A')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Í','I')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ô','O')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ô','O')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('É','E')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ú','U')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ê','E')
                                        primeiro_nome = primeiro_nome.upper()
                                        primeiro_nome = primeiro_nome.replace('Ç','C')
                                        primeiro_nome = primeiro_nome.replace('Ã','A')
                                        primeiro_nome = primeiro_nome.replace('À','A')
                                        primeiro_nome = primeiro_nome.replace('Á','A')
                                        primeiro_nome = primeiro_nome.replace('Í','I')
                                        primeiro_nome = primeiro_nome.replace('Ô','O')
                                        primeiro_nome = primeiro_nome.replace('Õ','O')
                                        primeiro_nome = primeiro_nome.replace('É','E')      
                                        primeiro_nome = primeiro_nome.replace('Ú','U')    
                                        primeiro_nome = primeiro_nome.replace('Ê','E')           
                                        print("Primeiro nome esperado: " + primeiro_nome)
                                        print("Primeiro nome encontrado na tela: " + primeiro_nome_na_tela)
                                        if primeiro_nome_na_tela.upper() == primeiro_nome.upper():
                                            print("Validação: Esperado " + primeiro_nome + " , encontrado " + primeiro_nome_na_tela + ".")
                                            print("[SUCESSO] Validação confirmada!")
                                            # Clicamos na linha inteira para selecionar
                                            codigo.click()
                                            print("Credor selecionado com sucesso.")
                                            
                                        else:
                                            print("[ERRO DE VALIDAÇÃO] O nome não corresponde ao esperado!")
                                            raise Exception('Validação falhou: Esperado ' + primeiro_nome + " , encontrado " + primeiro_nome_na_tela + ".")
                                        
                                except Exception as e:
                                        print(f"Ocorreu um erro durante a validação do credor: {e}")
                            
                            except:
                                print("Não encontrei o CPF")
                                todos_os_textos = codigo.all_inner_texts()
                                numeros_pc = []
                                    
                                if not numeros_pc:
                                    raise Exception("Nenhum número de CPF válido foi encontrado na lista de células.")
                    
                    manter_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    time.sleep(0.2)
                    valor_documento.press_sequentially(valor)
                    time.sleep(0.2)
                    observacao.press_sequentially(texto_da_ce)
                    time.sleep(1)
                    botao_incluir = manter_despesa_certificada.get_by_role("button", name="Incluir o Registro")
                    if robo_deve_parar:
                        if book:
                            print("Garantindo que a planilha seja fechada...")
                            book.close()
                        pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                        sys.exit()
                    
                    botao_incluir.click()

                    time.sleep(0.5)
                    manter_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    
                    try: 
                        print("Verificando se a mensagem 'Documento já cadastrado' existe...")
                        erro_na_tela = manter_despesa_certificada.get_by_role("cell", name="Número Documento já cadastrado(a).", exact=True)
                        if erro_na_tela.is_visible():
                            documento_ja_cadastrado = True
                        if documento_ja_cadastrado:
                            print("\n[AVISO] O documento já foi cadastrado anteriormente.")
                            print("O robô vai pular este item ou tomar uma ação alternativa.")
                        else:
                            print("\n[SUCESSO] Nenhuma mensagem de erro encontrada.")
                            print("Continuando com o fluxo normal da automação...")
                            documento_ja_cadastrado = False
                    except Exception as e:
                        print(f"Ocorreu um erro durante a verificação do documento: {e}")
                        documento_ja_cadastrado = False

                    if documento_ja_cadastrado == True:

                        despesa_certificada = "pesquisar no sigef"
                        pagina1.delete_rows(linha,1)
                        pagina1.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada])
                        book.save('Pagamentos.xlsx')   
                        
                    else:

                        
                        numero_despesa_certificada = manter_despesa_certificada.locator("#txtNuSeq")
                        numero_despesa_certificada.wait_for(timeout=10000)
                        numero_despesa_certificada.dblclick()
                        numero_despesa_certificada.press('Control+KeyC')
                        despesa_certificada =  numero_despesa_certificada = pyperclip.paste()
                        despesa_certificada = "2025CE" + str(despesa_certificada)
                        print(despesa_certificada)
                    
                        if despesa_certificada_teste == despesa_certificada:
                            print("REPETIDO. Refazendo Despesa Certificada!")
                            pagina1.delete_rows(linha,1)
                            book.save('Pagamentos.xlsx')
                            despesa_certificada = 'None'
                        else:
                            if despesa_certificada == 'despesa_certificada':
                                print("DEU ALGUM ERRO. Refazendo Despesa Certificada!")
                                pagina1.delete_rows(linha,1)
                                book.save('Pagamentos.xlsx')
                                despesa_certificada = 'None'
                            else:
                                if despesa_certificada == "pesquisar no sigef":
                                    time.sleep(0.1)
                                else:
                                    pagina1.delete_rows(linha,1)
                                    pagina1.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada])
                                    book.save('Pagamentos.xlsx')
                    
                    botao_limpar = manter_despesa_certificada.get_by_role("link", name="Limpar a Tela")
                    botao_limpar.click()
                    time.sleep(0.3)

            else:
                ja_foi_certificado = True

            if despesa_certificada != 'None':
                despesa_certificada_teste = despesa_certificada 
                linha = linha + 1
                documento_ja_cadastrado = False
 
        print("Nenhuma despesa para certificar. Iniciando a liquidação...")

        ja_foi_certificado = True
        manter_despesa_certificada.close()

#################################################################
#A PARTIR DAQUI COMEÇA A LIQUIDAR
#################################################################

        #PLANILHA NO EXCEL:
        try:
            book = openpyxl.load_workbook('Pagamentos.xlsx')
            pagina = book['Entrada']
            pagina1 = book['Despesas Certificadas']
            pagina2 = book['Notas de Liquidação']
            pagina3 = book['Preparações de Pagamento']
            pagina4 = book['Ordens Bancárias']
            pagina5 = book['Saída']
        except: 
            pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
            sys.exit()

        coluna = 1
        linha = 2

        while ja_foi_certificado == True:

            coluna = 1

            #LENDO A UG
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ug = cell.value
                    ug= str(ug)

            #LENDO A GESTÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    gestao = cell.value
                    gestao = str(gestao)

            #LENDO O Nº DO PROCESSO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    processo = cell.value
                    processo = str(processo)
                
            #LENDO O NOME DO SERVIDOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    nome = cell.value
                    nome = str(nome)

            #LENDO O CPF
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    cpf = cell.value
                    cpf = str(cpf)
                        
            #LENDO O VALOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    valor = cell.value
                    valor = str(valor)
                    valor = valor.replace('.',',')
                
            #LENDO O BANCO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    banco = cell.value
                    banco = str(banco)

            #LENDO A AGENCIA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    agencia = cell.value
                    agencia = str(agencia)

            #LENDO A CONTA CORRENTE
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    conta = cell.value
                    conta = str(conta)

            #LENDO A NOTA DE EMPENHO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    empenho = cell.value
                    empenho = str(empenho)

            #LENDO A DESPESA CERTIFICADA
            coluna = coluna + 1
            for row in pagina1.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    despesa_certificada = cell.value
                    despesa_certificada = str(despesa_certificada)

            #LENDO A NOTA DE LIQUIDAÇÃO
            coluna = coluna + 1
            for row in pagina2.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    liquidacao = cell.value
                    liquidacao = str(liquidacao)
                
            #LENDO A PREPARAÇÃO DE PAGAMENTO
            coluna = coluna + 1
            for row in pagina3.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    preparacao_pagamento = cell.value
                    preparacao_pagamento = str(preparacao_pagamento)

            #LENDO A ORDEM BANCÁRIA
            coluna = coluna + 1
            for row in pagina4.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ordem_bancaria = cell.value
                    ordem_bancaria = str(ordem_bancaria)

            #LENDO A DATA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    data = cell.value
                    data = str(data)

            #LENDO A OPERAÇÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    operacao = cell.value
                    operacao = str(operacao)

            if nome != "None":
                print("Estou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".")

            else:
                time.sleep(0)
                
            if nome == "None":
                if valor == "None":
                    ja_foi_certificado = False
                    break

            if cpf != "None":
                try:
                    if isinstance(cpf,str):
                        cpf = cpf.replace('.','')
                        cpf = cpf.replace('-','')
                        cpf_sem_ponto_virgula = int(cpf)
                    else: 
                        print('CPF é inválido.')
    
                    cpf_formatado = re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', "{:011d}".format(int(cpf_sem_ponto_virgula)))                        
                    cpf = cpf_formatado
                except:
                    time.sleep(0)    

            if isinstance(processo,str):
                processo = processo.replace('.','')
                processo = processo.replace('-','')
                processo = processo.replace('/','')
                processo_sem_pontos = int(processo)
            else: 
                print('Processo é inválido.')
            
            processo_formatado = re.sub(r'(\d{4})(\d{6})(\d{4})(\d{2})', r'\1.\2/\3-\4', "{:016d}".format(int(processo_sem_pontos))) 
            
            try:
                exercicio = int(exercicio)
            except:
                try:
                    empenho = int(empenho)
                except:
                    exercicio = str(exercicio)
                    empenho = str(empenho)

            if isinstance(empenho,str):
                empenho = empenho.upper()
                exercicio = empenho.strip().split('NE')[0]
                empenho = empenho.strip().split('NE')[1]
                exercicio = int(exercicio)
                empenho = int(empenho)
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
            else: 
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
                empenho = nota_de_empenho.strip().split('NE')[1]
                empenho = int(empenho)

            if despesa_certificada == "None":
                ja_foi_certificado = False
            
            else:
                ja_foi_certificado = True
                
                if liquidacao == "None":
                    ja_foi_liquidado = False
                
                else:
                    ja_foi_liquidado = True
    
            if ja_foi_liquidado == False:
                texto_da_nl =  "Liquidação de Despesa: Pagamento para o(a) servidor(a) " + str(cpf_formatado) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo_formatado) + "."
                texto_da_nl = texto_da_nl.replace('ç','c')
                texto_da_nl = texto_da_nl.replace('ã','a')
                texto_da_nl = texto_da_nl.replace('á','a')
                texto_da_nl = texto_da_nl.replace('à','a')
                texto_da_nl = texto_da_nl.replace('í','i')
                texto_da_nl = texto_da_nl.replace('ô','o')
                texto_da_nl = texto_da_nl.replace('õ','o')
                texto_da_nl = texto_da_nl.replace('é','e')
                pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
                pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
                pesquisar_funcionalidades_sistema.press_sequentially("Liquidar Despesa Certificada")
                funcionalidade_sistema = frame.get_by_title("Liquidar Despesa Certificada")
                    
                with guia.expect_popup() as popup_info:
                    
                    funcionalidade_sistema.click()
                    liquidar_despesa_certificada = popup_info.value

                    if robo_deve_parar:
                        if book:
                            print("Garantindo que a planilha seja fechada...")
                            book.close()
                        pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                        sys.exit()
                    
                    liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    campo_unidade_gestora = liquidar_despesa_certificada.get_by_role("textbox", name="Secretaria de Estado de")
                    campo_unidade_gestora.wait_for()
                    campo_unidade_gestora.press_sequentially(ug)
                    campo_gestao = liquidar_despesa_certificada.locator("#txtCdGestao_SIGEFPesquisa")

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    campo_gestao.press_sequentially(gestao)
                    campo_despesa_certificada = liquidar_despesa_certificada.locator("#txtDespesaCertificadaNumero_SIGEFPesquisa")
                    ce = despesa_certificada.replace("2025CE","")
                    campo_despesa_certificada.press_sequentially(ce)
                    botao_pesquisar = liquidar_despesa_certificada.get_by_role("button", name="Confirmar a Pesquisa")
                    botao_pesquisar.click()
                    
                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    try:
                        print("Procurando pela célula da Nota de Liquidação (NL)...")
                    
                    except:
                        ja_foi_liquidado = False

                    if ja_foi_liquidado == False:

                        if robo_deve_parar:
                            verificar_panico_e_sair(book)
                        
                        data_vencimento = liquidar_despesa_certificada.locator("#txtDataVencimento_SIGEFData")
                        data_vencimento.press_sequentially('h')
                        adicionar = liquidar_despesa_certificada.get_by_role("button", name="Adicionar Documento")
                        adicionar.click()
                        liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                        data_referencia = liquidar_despesa_certificada.locator("#txtDtReferenciaId_SIGEFData")
                        data_referencia.wait_for()
                        data_referencia.press_sequentially('h')
                        ponto_interrogacao = liquidar_despesa_certificada.locator("#txtNotaEmpenhoNumeroId_lnkBtnPesquisa")

                        if robo_deve_parar:
                            if book:
                                verificar_panico_e_sair(book)
                        
                        with liquidar_despesa_certificada.expect_popup() as popup_info:
                                ponto_interrogacao.click()
                                if robo_deve_parar:
                                    verificar_panico_e_sair(book)
                                selecionar_empenho = popup_info.value
                                selecionar_empenho.wait_for_load_state('networkidle', timeout=30000)
                                preencher_empenho = selecionar_empenho.locator("#txtNotaEmpenhoNumero")
                                preencher_empenho.press_sequentially(str(empenho))
                                botao_confirmar = selecionar_empenho.get_by_role("button", name="Confirmar a Consulta")
                                botao_confirmar.click()
                                selecionar_empenho.wait_for_load_state('networkidle', timeout=30000)
                                nota_empenho = selecionar_empenho.get_by_role("cell", name=nota_de_empenho, exact=True)
                                nota_empenho.wait_for()
                                nota_empenho.click()
                                
                        liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                        valor_bruto = liquidar_despesa_certificada.locator("#txtValorBrutoId")
                        valor_bruto.wait_for()
                        valor = str(valor)
                        time.sleep(0.3)
                        valor_bruto.press_sequentially(valor)
                        time.sleep(0.3)
                        botao_retencoes = liquidar_despesa_certificada.get_by_role("button", name="Sugerir Retenções")
                        botao_retencoes.click()
                        nao_existem_retencoes = liquidar_despesa_certificada.get_by_text("Não existem sugestões para")
                        nao_existem_retencoes.wait_for()
                        valor_liquido = liquidar_despesa_certificada.locator("#txtValorLiquidoId")
                        value_valor_liquido = valor_liquido.input_value()
                        
                        if value_valor_liquido == valor:

                            liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                            
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)

                            historico = liquidar_despesa_certificada.locator("#txtHistorico")
                            historico.wait_for()
                            historico.press('Control+KeyA')
                            historico.press('Delete')
                            historico.press_sequentially(texto_da_nl)
                            botao_confirmacao = liquidar_despesa_certificada.locator("#menun4").get_by_role("link")
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                            
                            botao_confirmacao.click()

                            botao_confirmar = liquidar_despesa_certificada.get_by_role("button", name="Confirmar a Operação")
                            
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                            
                            try:
                                botao_confirmar.click()
                                print("Verificando a existência de mensagens de erro...")
                                erro = liquidar_despesa_certificada.get_by_text("Não é permitido liquidar da")
                                erro_esta_visivel = erro.is_visible()
                                if erro_esta_visivel:
                                    print("[AVISO] Mensagem de erro detectada: 'Não é permitido liquidar da'.")
                                    botao_voltar = liquidar_despesa_certificada.get_by_role("button", name="Retornar à Tela Anterior")
                                    botao_voltar.click()
                                    documento_ja_liquidado_mas_nao_salvo = True
                                else:
                                    documento_ja_liquidado_mas_nao_salvo = False
                                    print("[SUCESSO] Nenhuma mensagem de erro encontrada. Continuando o fluxo normal.")
                            except Exception as e:
                                print(f"Ocorreu um erro inesperado durante a liquidação: {e}")

                            try:
                                liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                                print("Procurando pela célula da Nota de Liquidação (NL)...")
                                padrao_da_nl = re.compile(r"^2025NL\d{6}$")
                                celula_nl = liquidar_despesa_certificada.get_by_role("cell", name=padrao_da_nl)
                                celula_nl.first.wait_for(state="visible", timeout=30000)

                                if celula_nl.count() > 0:
                                    primeira_nl = celula_nl.first
                                    liquidacao = primeira_nl.inner_text()
                                    print(f"[SUCESSO] Nota de Liquidação encontrada e copiada: '{liquidacao}'")
                                else:
                                    liquidacao = "ERRO"
                                    print("[AVISO] Nenhuma Nota de Liquidação foi encontrada na página.")

                            except Exception as e:
                                print(f"Ocorreu um erro ao tentar localizar a NL: {e}")
                                liquidacao = "ERRO"

                            if liquidacao != "ERRO":
                                liquidacao = str(liquidacao) 
                                print(f"Valor final da variável 'liquidacao': {liquidacao}")
                                pagina2.delete_rows(linha,1)
                                pagina2.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia, conta,nota_de_empenho,despesa_certificada,liquidacao])
                                pagina5.delete_rows(linha,1)
                                pagina5.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia, conta,nota_de_empenho,despesa_certificada,liquidacao])
                                book.save('Pagamentos.xlsx')
                        else:
                            print("Deu algum erro!!")
                    else:
                            liquidacao = str(primeira_nl)
                            pagina2.delete_rows(linha,1)
                            pagina2.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia, conta,nota_de_empenho,despesa_certificada,liquidacao])
                            pagina5.delete_rows(linha,1)
                            pagina5.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia, conta,nota_de_empenho,despesa_certificada,liquidacao])
                            book.save('Pagamentos.xlsx')

                liquidar_despesa_certificada.close()
            linha = linha + 1
            coluna = 1

print("Fim das liquidações.")
if book:
    book.close()
print("\nScript finalizado. A janela de depuracao permanece aberta.")
keyboard.remove_hotkey(tecla_de_panico) 

pyautogui.alert(text='Encerrei por aqui.', title='Fim', button='OK')

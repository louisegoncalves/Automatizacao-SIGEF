#OLÁ!
#PROCEDIMENTO: CERTIFICAR E LIQUIDAR;
#POR: LOUISE-SESDEC;
#ALTERAÇÕES NO CÓDIGO PODEM SER ACESSADAS NO MEU GITHUB: <https://github.com/louisegoncalves/Automatizacao-SIGEF>.

#INSTRUÇÕES
#ATENÇÃO: É OBRIGATÓRIO ABRIR O DEPURADOR DO GOOGLE CHROME PARA EXECUTAR ESSE CÓDIGO
#EXECUTE NO CMD: "C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebugProfile"
#E LOGUE NO SIGEF

#BIBLIOTECAS UTILIZADAS:
from playwright.sync_api import sync_playwright, Page, TimeoutError
import pyautogui
import openpyxl
import keyboard
import time
import pyperclip
import re
from datetime import date
from datetime import datetime
import sys

#QUAL PLANILHA VAI SER UTILIZADA?
#planilha = "Pagamentos - voluntariar novembro.xlsx"
planilha = "Pagamentos.xlsx"

#VARIÁVEIS IMPORTANTES
robo_deve_parar = False
coluna = 1
linha = 2
loop = True
despesa_certificada_teste = 'None'

#PLANILHA NO EXCEL:
try:
    book = openpyxl.load_workbook(planilha)
    pagina = book['Entrada']
    pagina1 = book['Despesas Certificadas']
    pagina2 = book['Notas de Liquidação']
    pagina3 = book['Preparações de Pagamento']
    pagina4 = book['Ordens Bancárias']
    pagina5 = book['Saída']
except: 
    pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
    sys.exit()

#PLANILHA DE BACKUP:
try:
    book1 = openpyxl.load_workbook("Pagamentos_Backup.xlsx")
    pagina_backup = book1['Entrada']
    pagina1_backup = book1['Despesas Certificadas']
    pagina2_backup = book1['Notas de Liquidação']
    pagina3_backup = book1['Preparações de Pagamento']
    pagina4_backup = book1['Ordens Bancárias']
    pagina5_backup = book1['Saída']
except:
    try:
        wb = openpyxl.Workbook()
        ws_principal = wb.active
        ws_principal.title = "Entrada"
        wb.save("Pagamentos_Backup.xlsx")
        try:
            wb.create_sheet("Despesas Certificadas")
            wb.create_sheet("Notas de Liquidação")
            wb.create_sheet("Preparações de Pagamento")
            wb.create_sheet("Ordens Bancárias")
            wb.create_sheet('Saída')
            wb.save("Pagamentos_Backup.xlsx")
            print("Arquivo 'Pagamentos_Backup.xlsx' criado com sucesso com várias planilhas.")
        except:
            print("Erro na planilha de backup.")
        try:
            book1 = openpyxl.load_workbook("Pagamentos_Backup.xlsx")
            pagina_backup = book1['Entrada']
            pagina1_backup = book1['Despesas Certificadas']
            pagina2_backup = book1['Notas de Liquidação']
            pagina3_backup = book1['Preparações de Pagamento']
            pagina4_backup = book1['Ordens Bancárias']
            pagina5_backup = book1['Saída']
        except:
                print("Erro na planilha de backup.")
    except:
        print("Erro na planilha de backup.")

#SE QUISER DESATIVAR AQUELA JANELA DO COMEÇO PODE EXCLUIR ELA AQUI:
pyautogui.alert(text='Procedimento: Certificar, Liquidar e Preparar Pagamento.', title='Início', button='OK')

#FUNÇÃO QUE SERÁ CHAMADA PELA TECLA DE PANICO
def parar_execucao():
    global robo_deve_parar
    print("\n!!! TECLA ESC ACIONADA! ENCERRANDO AUTOMACAO !!!")
    robo_deve_parar = True

#FUNÇÃO QUE ENCERRA O CODIGO E FECHA A PLANILHA COM SEGURANÇA
#A PLANILHA DEVE SEMPRE SER FECHADA ANTES DE ENCERRAR, POIS CORRE O RISCO DE CORROMPER
def verificar_panico_e_sair(workbook):
    global robo_deve_parar
    if robo_deve_parar:
        print("Garantindo que a planilha seja fechada...")
        if workbook:
            workbook.close()
        pyautogui.alert('Tecla ESC acionada. Automação encerrada.')
        sys.exit()

#DEFINA SUA TECLA DE PÂNICO
tecla_de_panico = "Esc" 
keyboard.add_hotkey(tecla_de_panico, parar_execucao)
print(f"Robô iniciado. Pressione a tecla '{tecla_de_panico}' a qualquer momento para abortar com seguranca.")

#AQUI ELE VAI PEDIR PARA ABRIR O SIGEF PELO DEPURADOR DO GOOGLE
pyautogui.confirm(text='Aperte OK quando o SIGEF estiver logado no depurador do Google Chrome', title='Depurador do Chrome' , buttons=['OK'])

#PORTA DO DEPURADOR DO GOOGLE CHROME
CHROME_DEBUG_URL = "http://localhost:9222"

if robo_deve_parar:
    verificar_panico_e_sair(book)

#EXECUTANDO O PLAYWRIGHT DE FORMA SÍNCRONA
with sync_playwright() as p:
        if robo_deve_parar:
            verificar_panico_e_sair(book)

        #CONECTAR AO NAVEGADOR JÁ ABERTO:
        print(f"Tentando se conectar ao Chrome na porta de depuração: {CHROME_DEBUG_URL}")
        browser = p.chromium.connect_over_cdp(CHROME_DEBUG_URL)
        print("Conexão estabelecida com sucesso!")

        #OBTER A PÁGINA QUE ESTÁ ABERTA:
        #Quando conectamos, precisamos pegar o contexto e a página existentes;
        #Geralmente, a primeira página do primeiro contexto é a que queremos.
        janela = browser.contexts[0]
        guia = janela.pages[0]

        #VERIFICAR A PÁGINA ABERTA:
        print(f"Assumindo o controle da página com o título: '{guia.title()}'")
            
        #LOCALIZANDO O IFRAME:
        frame = guia.frame_locator('iframe[src="/SIGEF2025/SEG/#/SEGControleAcesso?p=1"]')
        
        #VERIFICAÇÃO DE PÂNICO:
        if robo_deve_parar:
            verificar_panico_e_sair(book)
        
        #INÍCIO                
        print("Iniciando!")

        #PESQUISANDO FUNCIONALIDADE NO SIGEF
        pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
        pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
        pesquisar_funcionalidades_sistema.press_sequentially("Manter Despesa Certificada")
        funcionalidade_sistema = frame.get_by_title("Manter Despesa Certificada")
        
        #ABRINDO JANELA MANTER DESPESA CERTIFICADA:
        with guia.expect_popup() as popup_info:
            funcionalidade_sistema.click()
            manter_despesa_certificada = popup_info.value
        
        #INÍCIO DO LOOP
        while loop == True:
            
            # A COLUNA SEMPRE DEVE REINICIAR NO INÍCIO DO LOOP
            coluna = 1
            
            #LENDO A UG
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ug = cell.value
                    ug= str(ug)

            #LENDO A GESTÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    gestao = cell.value
                    gestao = str(gestao)

            #LENDO O Nº DO PROCESSO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    processo = cell.value
                    processo = str(processo)
            
            #LENDO O NOME DO SERVIDOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    nome = cell.value
                    nome = str(nome)
                    primeiro_nome = nome.split()[0]

            #LENDO O CPF
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    cpf = cell.value
                    cpf = str(cpf)
                    
            #LENDO O VALOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):                
                for cell in row:
                    valor = cell.value
                    valor = str(valor)
            
            #LENDO O BANCO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    banco = cell.value
                    banco = str(banco)

            #LENDO A AGENCIA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    agencia = cell.value
                    agencia = str(agencia)

            #LENDO A CONTA CORRENTE
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    conta = cell.value
                    conta = str(conta)

            #LENDO A NOTA DE EMPENHO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    empenho = cell.value
                    empenho = str(empenho)

            #LENDO A DESPESA CERTIFICADA
            coluna = coluna + 1
            for row in pagina1.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    despesa_certificada = cell.value
                    despesa_certificada = str(despesa_certificada)

            #LENDO A NOTA DE LIQUIDAÇÃO
            coluna = coluna + 1
            for row in pagina2.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    liquidacao = cell.value
                    liquidacao = str(liquidacao)
            
            #LENDO A PREPARAÇÃO DE PAGAMENTO
            coluna = coluna + 1
            for row in pagina3.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    preparacao_pagamento = cell.value
                    preparacao_pagamento = str(preparacao_pagamento)

            #LENDO A ORDEM BANCÁRIA
            coluna = coluna + 1
            for row in pagina4.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ordem_bancaria = cell.value
                    ordem_bancaria = str(ordem_bancaria)

            #LENDO A DATA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    data = cell.value
                    data = str(data)

            #LENDO A OPERAÇÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    operacao = cell.value
                    operacao = str(operacao)
            
            #FORMATANDO O CAMPO DA PLANILHA "CPF" NO MOLDE 000.000.000-00
            #O RESULTADO SERÁ DUAS VARIÁVEIS: cpf_sem_ponto_virgula e cpf_formatado
            if cpf != "None":
                try:
                    if isinstance(cpf,str):
                        cpf = cpf.replace('.','')
                        cpf = cpf.replace('-','')
                        cpf_sem_ponto_virgula = str(cpf)
                    else: 
                        print('CPF é inválido.')
    
                    cpf_formatado = re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', "{:011d}".format(int(cpf_sem_ponto_virgula)))                        
                    cpf = cpf_formatado
                except:
                    time.sleep(0)    
                   
            if nome != "None":
                print("Estou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".")
            else:
                time.sleep(0)

            if nome == "None":
                if valor == "None":
                    loop = False
                    break

            if data != "None":
                if isinstance(data,str):
                    data_fomatada = data.replace('/','')
                    data_fomatada = data_fomatada.replace('.','')
                    data_fomatada = data_fomatada.replace('-','')
                    data_fomatada = data_fomatada.replace(' ','')
                    data_fomatada = data_fomatada.replace(',','')
                            
                    try:
                        data_formatada = int(data_fomatada)
                        data_formatada = re.sub(r'(\d{2})(\d{2})(\d{4})', r'\1/\2/\3', "{:08d}".format(int(data_fomatada)))
                        data_foi_formatada = True
                    except:
                        print('Atenção, na planilha consta que a data é ' + data + ", portanto será utilizado como parâmetro a data atual.")
                        data_atual = date.today() 
                        data_formatada = data_atual.strftime("%d/%m/%Y")
                        data_foi_formatada = True
                        
                else:
                    print('Atenção, na planilha consta que a data é ' + data + ", portanto será utilizado como parâmetro a data atual.")
                    data_atual = date.today() 
                    data_formatada = data_atual.strftime("%d/%m/%Y")
                    data_foi_formatada = True
                        
            else: 
                print('[ATENÇÃO] Data não foi preenchida na planilha.')
                data_foi_formatada = False
        
            if isinstance(processo,str):
                processo = processo.replace('.','')
                processo = processo.replace('-','')
                processo = processo.replace('/','')
                processo_sem_pontos = str(processo)

            else: 
                print('Processo é inválido.')
            
            processo_formatado = re.sub(r'(\d{4})(\d{6})(\d{4})(\d{2})', r'\1.\2/\3-\4', "{:016d}".format(int(processo_sem_pontos))) 
            #AQUI SELECIONAMOS O NÚMERO DO MEIO DO PROCESSO:
            processo_cortado = processo_formatado.strip().split('/')[0]
            processo_cortado = processo_cortado.strip().split('.')[1]
            linha_documento = int(linha) - 1
            value_numero_cortado = str(processo_cortado) + "-" + str(linha_documento)
        
            try:
                exercicio = int(exercicio)
            except:
                try:
                    empenho = int(empenho)
                except:
                    exercicio = '2025'
                    empenho = str(empenho)

            if isinstance(empenho,str):
                empenho = empenho.upper()
                exercicio = empenho.strip().split('NE')[0]
                empenho = empenho.strip().split('NE')[1]
                exercicio = int(exercicio)
                empenho = int(empenho)
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
            else: 
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)

            if despesa_certificada == "None":
                ja_foi_certificado = False
                if ja_foi_certificado == False:
                    texto_da_ce =  "Certificação de Despesa: Pagamento para o(a) servidor(a) " + str(cpf_formatado) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo_formatado) + "."

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    #INFORMAÇÕES PRELIMINARES
                    #DATA DE HOJE:
                    data_atual = date.today() 
                    #FORMATAR A DATA:
                    data_de_baixa = data_atual.strftime("%d/%m/%Y")
                    #HORA:
                    agora = datetime.now()
                    #FORMATAR A HORA:
                    hora = agora.strftime("%H:%M:%S")
                    manter_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    time.sleep(0.1)
                    campo_gestao = manter_despesa_certificada.locator("#txtCdGestao_SIGEFPesquisa")
                    campo_gestao.wait_for()
                    campo_gestao.press_sequentially(gestao)
                    tipo_documento = manter_despesa_certificada.locator("#cmbCdTipoDocumento")
                    tipo_documento.select_option(label="Outros")
                    numero_documento = manter_despesa_certificada.locator("#txtNuDocumento")
                    numero_documento.press_sequentially(value_numero_cortado)
                    favorecido = manter_despesa_certificada.locator("#txtNmCredor_lnkBtnPesquisa")
                    valor_documento = manter_despesa_certificada.locator("#txtVlDocumento")
                    data_emissao = manter_despesa_certificada.locator("#txtDtEmissao_SIGEFData")
                    data_aceite = manter_despesa_certificada.locator("#txtDtAceite_SIGEFData")
                    data_apresentacao = manter_despesa_certificada.locator("#txtDtApresentacao_SIGEFData")
                    competencia = manter_despesa_certificada.locator("#cboMesComp")
                    observacao = manter_despesa_certificada.locator("#txtDeObservacao")
                    atestado = manter_despesa_certificada.get_by_role("checkbox", name="Sou responsável pelo atesto")
                    time.sleep(0.2)
                    data_emissao.click()
                    data_emissao.press_sequentially(data_de_baixa)
                    time.sleep(0.2)
                    data_aceite.click()
                    data_aceite.press_sequentially(data_de_baixa)
                    time.sleep(0.2)
                    data_apresentacao.click()
                    data_apresentacao.press_sequentially(data_de_baixa)
                    time.sleep(0.2)

                    if data_foi_formatada == True:

                        mes = data_formatada.strip().split('/')[1]
                        mes = mes.strip().split('/')[0]
                        print(mes)

                        if mes == "01":
                            selecionar_competencia = 'Janeiro'
                        else:
                            if mes == "02":
                                selecionar_competencia = 'Fevereiro'
                            else:
                                if mes == "03":
                                    selecionar_competencia = 'Março'
                                else:
                                    if mes == "04":
                                        selecionar_competencia = 'Abril'
                                    else:
                                        if mes == "05":
                                            selecionar_competencia = 'Maio'
                                        else:
                                            if mes == "06":
                                                selecionar_competencia = 'Junho'
                                            else:
                                                if mes == "07":
                                                    selecionar_competencia = 'Julho'
                                                else:   
                                                    if mes == "08":
                                                        selecionar_competencia = 'Agosto'
                                                    else:
                                                        if mes == "09":
                                                            selecionar_competencia = 'Setembro'
                                                        else:
                                                            if mes == "10":
                                                                selecionar_competencia = 'Outubro'
                                                            else:
                                                                if mes == "11":
                                                                    selecionar_competencia = 'Novembro'
                                                                else:
                                                                    if mes == "12":
                                                                        selecionar_competencia = 'Dezembro'
                                                                    else:
                                                                        selecionar_competencia = 'Setembro'
                        data_foi_formatada = False
                    
                    competencia.select_option(label=selecionar_competencia)
                    time.sleep(0.2)
                    
                    atestado.click()
                    
                    with manter_despesa_certificada.expect_popup() as popup_info:
                            favorecido.click()
                            selecionar_favorecido = popup_info.value

                            if robo_deve_parar:
                                pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                                
                            selecionar_favorecido.wait_for_load_state('networkidle', timeout=30000)
                            time.sleep(0.5)
                            botao_cpf = selecionar_favorecido.locator("#btnCPF")
                            botao_cpf.click()
                            preencher_cpf = selecionar_favorecido.get_by_role("textbox")
                            preencher_cpf.wait_for()
                            cpf_sem_ponto_virgula = str(cpf_sem_ponto_virgula)
                            preencher_cpf.press_sequentially(cpf_sem_ponto_virgula)
                            time.sleep(0.5)
                            botao_confirmar = selecionar_favorecido.get_by_role("button", name="Confirmar a Consulta")
                            botao_confirmar.click()
                            selecionar_favorecido.wait_for_load_state('networkidle', timeout=30000)
                            localizar_funcao = selecionar_favorecido.get_by_text("* CPF")
                            localizar_funcao.wait_for()
                            try:
                                codigo = selecionar_favorecido.get_by_role("cell", name=cpf_formatado, exact=True)
                                codigo.wait_for()
                                
                                try:
                                    padrao_cpf = re.compile(r"\d{3}\.\d{3}\.\d{3}-\d{2}")
    
                                    #AQUI ENCONTRAMOS A PRIMEIRA CÉLULA QUE CORRESPONDE AO PADRÃO DO CPF:
                                    primeira_celula_cpf = selecionar_favorecido.get_by_text(padrao_cpf).first
                                        
                                    #É crucial esperar que esta âncora apareça
                                    primeira_celula_cpf.wait_for(timeout=10000)
                                    print("Célula âncora encontrada com o texto: " + primeira_celula_cpf.inner_text())

                                    # --- PASSO 2: A PARTIR DA ÂNCORA, NAVEGAR ATÉ A LINHA PAI ---
                                    # O XPath '..' significa "vá para o elemento pai".
                                    # O pai de uma célula (<td>) é a sua linha (<tr>).
                                    linha_correta = primeira_celula_cpf.locator("xpath=..")
                                        
                                    # --- PASSO 3: DA LINHA, NAVEGAR ATÉ A CÉLULA DO NOME ---
                                    # Agora que temos a linha correta, pegamos a segunda célula (índice 1)
                                    celula_nome_credor = linha_correta.locator("td").nth(1)
                                        
                                    nome_completo_na_tela = celula_nome_credor.inner_text()
                                    nome_completo_na_tela = nome_completo_na_tela.upper()
                                        
                                    # --- PASSO 4: VALIDAR E AGIR ---
                                    primeiro_nome_na_tela = ""
                                    if nome_completo_na_tela and nome_completo_na_tela.strip():
                                        primeiro_nome_na_tela = nome_completo_na_tela.strip().split()[0]
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ç','C')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ã','A')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Á','A')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('À','A')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Í','I')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ô','O')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ô','O')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('É','E')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ú','U')
                                        primeiro_nome_na_tela = primeiro_nome_na_tela.replace('Ê','E')
                                        primeiro_nome = primeiro_nome.upper()
                                        primeiro_nome = primeiro_nome.replace('Ç','C')
                                        primeiro_nome = primeiro_nome.replace('Ã','A')
                                        primeiro_nome = primeiro_nome.replace('À','A')
                                        primeiro_nome = primeiro_nome.replace('Â','A')
                                        primeiro_nome = primeiro_nome.replace('Á','A')
                                        primeiro_nome = primeiro_nome.replace('Í','I')
                                        primeiro_nome = primeiro_nome.replace('Ô','O')
                                        primeiro_nome = primeiro_nome.replace('Õ','O')
                                        primeiro_nome = primeiro_nome.replace('É','E')      
                                        primeiro_nome = primeiro_nome.replace('Ú','U')    
                                        primeiro_nome = primeiro_nome.replace('Ê','E')           
                                        print("Primeiro nome esperado: " + primeiro_nome)
                                        print("Primeiro nome encontrado na tela: " + primeiro_nome_na_tela)
                                        if primeiro_nome_na_tela.upper() == primeiro_nome.upper():
                                            print("Validação: Esperado " + primeiro_nome + " , encontrado " + primeiro_nome_na_tela + ".")
                                            print("[SUCESSO] Validação confirmada!")
                                            # Clicamos na linha inteira para selecionar
                                            codigo.click()
                                            print("Credor selecionado com sucesso.")
                                            
                                        else:
                                            print("[ERRO DE VALIDAÇÃO] O nome não corresponde ao esperado!")
                                            raise Exception('Validação falhou: Esperado ' + primeiro_nome + " , encontrado " + primeiro_nome_na_tela + ".")
                                        
                                except Exception as e:
                                        print(f"Ocorreu um erro durante a validação do credor: {e}")
                            
                            except:
                                print("Não encontrei o CPF")
                                todos_os_textos = codigo.all_inner_texts()
                                numeros_pc = []
                                    
                                if not numeros_pc:
                                    raise Exception("Nenhum número de CPF válido foi encontrado na lista de células.")
                    
                    manter_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    time.sleep(0.2)
                    valor_documento.press_sequentially(valor)
                    time.sleep(0.2)
                    observacao.press_sequentially(texto_da_ce)
                    time.sleep(1)
                    botao_incluir = manter_despesa_certificada.get_by_role("button", name="Incluir o Registro")
                    if robo_deve_parar:
                        if book:
                            print("Garantindo que a planilha seja fechada...")
                            book.close()
                        pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                        sys.exit()
                    
                    botao_incluir.click()

                    time.sleep(0.5)
                    manter_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    
                    try: 
                        print("Verificando se a mensagem 'Documento já cadastrado' existe...")
                        erro_na_tela = manter_despesa_certificada.get_by_role("cell", name="Número Documento já cadastrado(a).", exact=True)
                        if erro_na_tela.is_visible():
                            documento_ja_cadastrado = True
                        if documento_ja_cadastrado:
                            print("\n[AVISO] O documento já foi cadastrado anteriormente.")
                            print("O robô vai pular este item ou tomar uma ação alternativa.")
                        else:
                            print("\n[SUCESSO] Nenhuma mensagem de erro encontrada.")
                            print("Continuando com o fluxo normal da automação...")
                            documento_ja_cadastrado = False
                    except Exception as e:
                        print(f"Ocorreu um erro durante a verificação do documento: {e}")
                        documento_ja_cadastrado = False

                    if documento_ja_cadastrado == True:
                        try:
                            despesa_certificada = "pesquisar no sigef"
                            pagina1_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada])
                            pagina1.delete_rows(linha,1)
                            pagina1.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada])
                            book.save(planilha)   
                        except:
                            book1.save("Pagamentos_Backup.xlsx")
                            print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                            book1.close()
                            sys.exit()
                    else:
                        numero_despesa_certificada = manter_despesa_certificada.locator("#txtNuSeq")
                        numero_despesa_certificada.wait_for(timeout=10000)
                        numero_despesa_certificada.dblclick()
                        numero_despesa_certificada.press('Control+KeyC')
                        despesa_certificada =  numero_despesa_certificada = pyperclip.paste()
                        despesa_certificada = "2025CE" + str(despesa_certificada)
                        print(despesa_certificada)
                    
                        if despesa_certificada_teste == despesa_certificada:
                            print("REPETIDO. Refazendo Despesa Certificada!")
                            pagina1.delete_rows(linha,1)
                            book.save(planilha)
                            despesa_certificada = 'None'
                        else:
                            if despesa_certificada == 'despesa_certificada':
                                print("DEU ALGUM ERRO. Refazendo Despesa Certificada!")
                                pagina1.delete_rows(linha,1)
                                book.save(planilha)
                                despesa_certificada = 'None'
                            else:
                                if despesa_certificada == "pesquisar no sigef":
                                    time.sleep(0.1)
                                else:
                                    try:
                                        pagina1_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada])
                                        pagina1.delete_rows(linha,1)
                                        pagina1.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada])
                                        book.save(planilha)
                                    except:
                                        book1.save("Pagamentos_Backup.xlsx")
                                        print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                        book1.close()
                                        sys.exit()
                    
                    botao_limpar = manter_despesa_certificada.get_by_role("link", name="Limpar a Tela")
                    botao_limpar.click()
                    time.sleep(0.3)

            else:
                ja_foi_certificado = True

            if despesa_certificada != 'None':
                despesa_certificada_teste = despesa_certificada 
                linha = linha + 1
                documento_ja_cadastrado = False
 
        print("Nenhuma despesa para certificar. Iniciando a liquidação...")

        ja_foi_certificado = True
        manter_despesa_certificada.close()

#################################################################
#A PARTIR DAQUI COMEÇA A LIQUIDAR
#################################################################

        #PLANILHA NO EXCEL:
        try:
            book = openpyxl.load_workbook(planilha)
            pagina = book['Entrada']
            pagina1 = book['Despesas Certificadas']
            pagina2 = book['Notas de Liquidação']
            pagina3 = book['Preparações de Pagamento']
            pagina4 = book['Ordens Bancárias']
            pagina5 = book['Saída']
        except: 
            pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
            sys.exit()

        coluna = 1
        linha = 2

        pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
        pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
        pesquisar_funcionalidades_sistema.press_sequentially("Liquidar Despesa Certificada")
        funcionalidade_sistema = frame.get_by_title("Liquidar Despesa Certificada")
                    
        with guia.expect_popup() as popup_info:
                    
            funcionalidade_sistema.click()
            liquidar_despesa_certificada = popup_info.value

            while ja_foi_certificado == True:

                coluna = 1

                #LENDO A UG
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        ug = cell.value
                        ug= str(ug)

                #LENDO A GESTÃO
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        gestao = cell.value
                        gestao = str(gestao)

                #LENDO O Nº DO PROCESSO
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        processo = cell.value
                        processo = str(processo)
                    
                #LENDO O NOME DO SERVIDOR
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        nome = cell.value
                        nome = str(nome)

                #LENDO O CPF
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        cpf = cell.value
                        cpf = str(cpf)
                            
                #LENDO O VALOR
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        valor = cell.value
                        valor = str(valor)
                    
                #LENDO O BANCO
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        banco = cell.value
                        banco = str(banco)

                #LENDO A AGENCIA
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        agencia = cell.value
                        agencia = str(agencia)

                #LENDO A CONTA CORRENTE
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        conta = cell.value
                        conta = str(conta)

                #LENDO A NOTA DE EMPENHO
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        empenho = cell.value
                        empenho = str(empenho)

                #LENDO A DESPESA CERTIFICADA
                coluna = coluna + 1
                for row in pagina1.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        despesa_certificada = cell.value
                        despesa_certificada = str(despesa_certificada)

                #LENDO A NOTA DE LIQUIDAÇÃO
                coluna = coluna + 1
                for row in pagina2.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        liquidacao = cell.value
                        liquidacao = str(liquidacao)
                    
                #LENDO A PREPARAÇÃO DE PAGAMENTO
                coluna = coluna + 1
                for row in pagina3.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        preparacao_pagamento = cell.value
                        preparacao_pagamento = str(preparacao_pagamento)

                #LENDO A ORDEM BANCÁRIA
                coluna = coluna + 1
                for row in pagina4.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        ordem_bancaria = cell.value
                        ordem_bancaria = str(ordem_bancaria)

                #LENDO A DATA
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        data = cell.value
                        data = str(data)

                #LENDO A OPERAÇÃO
                coluna = coluna + 1
                for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                    for cell in row:
                        operacao = cell.value
                        operacao = str(operacao)

                if nome != "None":
                    print("Estou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".")

                else:
                    time.sleep(0)
                    
                if nome == "None":
                    if valor == "None":
                        ja_foi_certificado = False
                        break

                if cpf != "None":
                    try:
                        if isinstance(cpf,str):
                            cpf = cpf.replace('.','')
                            cpf = cpf.replace('-','')
                            cpf_sem_ponto_virgula = int(cpf)
                        else: 
                            print('CPF é inválido.')
        
                        cpf_formatado = re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', "{:011d}".format(int(cpf_sem_ponto_virgula)))                        
                        cpf = cpf_formatado
                    except:
                        time.sleep(0)    

                if isinstance(processo,str):
                    processo = processo.replace('.','')
                    processo = processo.replace('-','')
                    processo = processo.replace('/','')
                    processo_sem_pontos = int(processo)
                else: 
                    print('Processo é inválido.')
                
                processo_formatado = re.sub(r'(\d{4})(\d{6})(\d{4})(\d{2})', r'\1.\2/\3-\4', "{:016d}".format(int(processo_sem_pontos))) 
                
                try:
                    exercicio = int(exercicio)
                except:
                    try:
                        empenho = int(empenho)
                    except:
                        exercicio = str(exercicio)
                        empenho = str(empenho)

                if isinstance(empenho,str):
                    empenho = empenho.upper()
                    exercicio = empenho.strip().split('NE')[0]
                    empenho = empenho.strip().split('NE')[1]
                    exercicio = int(exercicio)
                    empenho = int(empenho)
                    exercicio_NE = str(exercicio) + "NE"
                    nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                    nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
                else: 
                    exercicio_NE = str(exercicio) + "NE"
                    nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                    nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
                    empenho = nota_de_empenho.strip().split('NE')[1]
                    empenho = int(empenho)

                if despesa_certificada == "None":
                    ja_foi_certificado = False
                
                else:
                    ja_foi_certificado = True
                    
                    if liquidacao == "None":
                        ja_foi_liquidado = False
                    
                    else:
                        ja_foi_liquidado = True
        
                if ja_foi_liquidado == False:
                    texto_da_nl =  "Liquidação de Despesa: Pagamento para o(a) servidor(a) " + str(cpf_formatado) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo_formatado) + "."


                    if robo_deve_parar:
                        if book:
                            print("Garantindo que a planilha seja fechada...")
                            book.close()
                        pyautogui.alert(text='Tecla ESC acionada. Automacao encerrada', title='Tecla de Panico Acionada', button='OK')
                        sys.exit()
                        
                    liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                    campo_unidade_gestora = liquidar_despesa_certificada.get_by_role("textbox", name="Secretaria de Estado de")
                    campo_unidade_gestora.wait_for()
                    campo_unidade_gestora.press_sequentially(ug)
                    campo_gestao = liquidar_despesa_certificada.locator("#txtCdGestao_SIGEFPesquisa")

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    campo_gestao.press_sequentially(gestao)
                    campo_despesa_certificada = liquidar_despesa_certificada.locator("#txtDespesaCertificadaNumero_SIGEFPesquisa")
                    ce = despesa_certificada.replace("2025CE","")
                    campo_despesa_certificada.press_sequentially(ce)
                    botao_pesquisar = liquidar_despesa_certificada.get_by_role("button", name="Confirmar a Pesquisa")
                    botao_pesquisar.click()
                        
                    if robo_deve_parar:
                            verificar_panico_e_sair(book)

                    try:
                            print("Procurando pela célula da Nota de Liquidação (NL)...")
                        
                    except:
                        ja_foi_liquidado = False

                    if ja_foi_liquidado == False:

                        if robo_deve_parar:
                            verificar_panico_e_sair(book)
                            
                        data_vencimento = liquidar_despesa_certificada.locator("#txtDataVencimento_SIGEFData")
                        data_vencimento.press_sequentially('h')
                        adicionar = liquidar_despesa_certificada.get_by_role("button", name="Adicionar Documento")
                        adicionar.click()
                        liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                        data_referencia = liquidar_despesa_certificada.locator("#txtDtReferenciaId_SIGEFData")
                        data_referencia.wait_for()
                        data_referencia.press_sequentially('h')
                        ponto_interrogacao = liquidar_despesa_certificada.locator("#txtNotaEmpenhoNumeroId_lnkBtnPesquisa")

                        if robo_deve_parar:
                            if book:
                                verificar_panico_e_sair(book)
                            
                        with liquidar_despesa_certificada.expect_popup() as popup_info:
                                ponto_interrogacao.click()
                                if robo_deve_parar:
                                    verificar_panico_e_sair(book)
                                selecionar_empenho = popup_info.value
                                selecionar_empenho.wait_for_load_state('networkidle', timeout=30000)
                                preencher_empenho = selecionar_empenho.locator("#txtNotaEmpenhoNumero")
                                preencher_empenho.press_sequentially(str(empenho))
                                botao_confirmar = selecionar_empenho.get_by_role("button", name="Confirmar a Consulta")
                                botao_confirmar.click()
                                selecionar_empenho.wait_for_load_state('networkidle', timeout=30000)
                                nota_empenho = selecionar_empenho.get_by_role("cell", name=nota_de_empenho, exact=True)
                                nota_empenho.wait_for()
                                nota_empenho.click()
                                    
                        liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                        valor_bruto = liquidar_despesa_certificada.locator("#txtValorBrutoId")
                        valor_bruto.wait_for()
                        valor = str(valor)
                        try:
                            time.sleep(0.3)
                            valor_bruto.press_sequentially(valor)
                        except:
                            try:
                                time.sleep(0.3)
                                valor_bruto.press_sequentially(valor)
                            except:
                                texto = "Insira manualmente o valor de " + valor + '.'
                                pyautogui.alert(text=texto, title='Fim', button='OK')

                        time.sleep(0.3)
                        botao_retencoes = liquidar_despesa_certificada.get_by_role("button", name="Sugerir Retenções")
                        botao_retencoes.click()
                        nao_existem_retencoes = liquidar_despesa_certificada.get_by_text("Não existem sugestões para")
                        nao_existem_retencoes.wait_for()
                        valor_liquido = liquidar_despesa_certificada.locator("#txtValorLiquidoId")
                        valor_liquido.wait_for()
                        value_valor_liquido = valor_liquido.input_value()
                        
                        try:
                            value_valor_liquido = value_valor_liquido.replace('.','')
                        except:
                            time.sleep(0)
                        try:
                            valor = valor.replace('.','')
                        except:
                            time.sleep(0)
                        try:
                            valor = valor.replace(' ','')
                        except:
                            time.sleep(0)
                        
                            
                        if value_valor_liquido == valor:

                            liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                                
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)

                            historico = liquidar_despesa_certificada.locator("#txtHistorico")
                            historico.wait_for()
                            historico.press('Control+KeyA')
                            historico.press('Delete')
                            historico.press_sequentially(texto_da_nl)
                            botao_confirmacao = liquidar_despesa_certificada.locator("#menun4").get_by_role("link")
                            
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                                
                            botao_confirmacao.click()

                            botao_confirmar = liquidar_despesa_certificada.get_by_role("button", name="Confirmar a Operação")
                            botao_limpar = liquidar_despesa_certificada.get_by_role("link", name="Limpar a Tela")
                                
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                                
                            try:
                                botao_confirmar.click()
                                print("Verificando a existência de mensagens de erro...")
                                try:
                                    erro = liquidar_despesa_certificada.get_by_text("Não é permitido liquidar da")
                                    erro_esta_visivel = erro.is_visible()
                                    if erro_esta_visivel:
                                        print("[AVISO] Mensagem de erro detectada: 'Não é permitido liquidar da'.")
                                        botao_voltar = liquidar_despesa_certificada.get_by_role("button", name="Retornar à Tela Anterior")
                                        botao_voltar.click()
                                        documento_ja_liquidado_mas_nao_salvo = True
                                    else:
                                        documento_ja_liquidado_mas_nao_salvo = False
                                        print("[SUCESSO] Nenhuma mensagem de erro encontrada. Continuando o fluxo normal.")
                                except:
                                        time.sleep(1)
                            except Exception as e:
                                print(f"Ocorreu um erro inesperado durante a liquidação: {e}")

                            try:
                                liquidar_despesa_certificada.wait_for_load_state('networkidle', timeout=30000)
                                print("Procurando pela célula da Nota de Liquidação (NL)...")
                                padrao_da_nl = re.compile(r"^2025NL\d{6}$")
                                celula_nl = liquidar_despesa_certificada.get_by_role("cell", name=padrao_da_nl)
                                celula_nl.first.wait_for(state="visible", timeout=30000)

                                if celula_nl.count() > 0:
                                        primeira_nl = celula_nl.first
                                        liquidacao = primeira_nl.inner_text()
                                        print(f"[SUCESSO] Nota de Liquidação encontrada e copiada: '{liquidacao}'")
                                else:
                                        liquidacao = "ERRO"
                                        print("[AVISO] Nenhuma Nota de Liquidação foi encontrada na página.")

                            except Exception as e:
                                    print(f"Ocorreu um erro ao tentar localizar a NL: {e}")
                                    liquidacao = "ERRO"

                            if liquidacao != "ERRO":
                                liquidacao = str(liquidacao) 
                                print(f"Valor final da variável 'liquidacao': {liquidacao}")
                                try:
                                        pagina2_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada,liquidacao])
                                        pagina2.delete_rows(linha,1)
                                        pagina2.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia, conta,nota_de_empenho,despesa_certificada,liquidacao])
                                        book.save(planilha)
                                except:           
                                        
                                        book1.save("Pagamentos_Backup.xlsx")
                                        print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                        book1.close()
                                        sys.exit()

                        else:

                                print("Deu algum erro!!")
                    else:
                        
                        liquidacao = str(primeira_nl)
                        try:
                                    pagina2_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada,liquidacao])
                                    pagina2.delete_rows(linha,1)
                                    pagina2.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia, conta,nota_de_empenho,despesa_certificada,liquidacao])
                                    book.save(planilha)
                        except:
                                                                    
                                    book1.save("Pagamentos_Backup.xlsx")
                                    print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                    book1.close()
                                    sys.exit()
                try:         
                    botao_limpar.click()
                except:
                    time.sleep(0)
                    
                linha = linha + 1
                coluna = 1


        try:
            liquidar_despesa_certificada.close()
        except:
            time.sleep(0)

        print("Fim das liquidações.")

        if book:
            book.close()

        #print("\nScript finalizado. A janela de depuracao permanece aberta.")
        #keyboard.remove_hotkey(tecla_de_panico) 
        #pyautogui.alert(text='Encerrei por aqui.', title='Fim', button='OK')


#################################################################
#A PARTIR DAQUI COMEÇA A PREPARAR PAGAMENTO
#################################################################

        #PLANILHA NO EXCEL:
        try:
            book = openpyxl.load_workbook(planilha)
            pagina = book['Entrada']
            pagina1 = book['Despesas Certificadas']
            pagina2 = book['Notas de Liquidação']
            pagina3 = book['Preparações de Pagamento']
            pagina4 = book['Ordens Bancárias']
            pagina5 = book['Saída']
        except: 
            pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
            sys.exit()

        coluna = 1
        linha = 2

        pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
        pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
        pesquisar_funcionalidades_sistema.press_sequentially("PP Despesa Empenhada")
        funcionalidade_sistema = frame.get_by_title("PP Despesa Empenhada")
        
        with guia.expect_popup() as popup_info:
            funcionalidade_sistema.click()
            pp_despesa_empenhada = popup_info.value
        
        preparando_pagamento = True

        while preparando_pagamento == True:
            
            coluna = 1
            

            #LENDO A UG
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ug = cell.value
                    ug= str(ug)

            #LENDO A GESTÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    gestao = cell.value
                    gestao = str(gestao)

            #LENDO O Nº DO PROCESSO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    processo = cell.value
                    processo = str(processo)
            
            #LENDO O NOME DO SERVIDOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    nome = cell.value
                    nome = str(nome)
                    primeiro_nome = nome.split()[0]

            #LENDO O CPF
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    cpf = cell.value
                    cpf = str(cpf)
                    
            #LENDO O VALOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):                
                for cell in row:
                    valor = cell.value
                    valor = str(valor)
                    valor = valor.replace('.',',')
            
            #LENDO O BANCO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    banco = cell.value
                    banco = str(banco)

            #LENDO A AGENCIA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    agencia = cell.value
                    agencia = str(agencia)

            #LENDO A CONTA CORRENTE
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    conta = cell.value
                    conta = str(conta)

            #LENDO A NOTA DE EMPENHO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    empenho = cell.value
                    empenho = str(empenho)

            #LENDO A DESPESA CERTIFICADA
            coluna = coluna + 1
            for row in pagina1.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    despesa_certificada = cell.value
                    despesa_certificada = str(despesa_certificada)

            #LENDO A NOTA DE LIQUIDAÇÃO
            coluna = coluna + 1
            for row in pagina2.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    liquidacao = cell.value
                    liquidacao = str(liquidacao)
            
            #LENDO A PREPARAÇÃO DE PAGAMENTO
            coluna = coluna + 1
            for row in pagina3.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    preparacao_pagamento = cell.value
                    preparacao_pagamento = str(preparacao_pagamento)

            #LENDO A ORDEM BANCÁRIA
            coluna = coluna + 1
            for row in pagina4.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ordem_bancaria = cell.value
                    ordem_bancaria = str(ordem_bancaria)

            #LENDO A DATA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    data = cell.value
                    data = str(data)

            #LENDO A OPERAÇÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    operacao = cell.value
                    operacao = str(operacao)
            
            if liquidacao != None:
                ja_foi_liquidado = True
                if preparacao_pagamento != "None":
                    ja_foi_preparado = True
                else:
                    ja_foi_preparado = False
            else:
                ja_foi_liquidado = False
                preparando_pagamento = False

            if robo_deve_parar:
                verificar_panico_e_sair(book)
            
            if cpf != "None":
                try:
                    if isinstance(cpf,str):
                        cpf = cpf.replace('.','')
                        cpf = cpf.replace('-','')
                        cpf_sem_ponto_virgula = int(cpf)
                    else: 
                        print('CPF é inválido.')
                    
                    cpf_formatado = re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', "{:011d}".format(int(cpf_sem_ponto_virgula)))                        
                    cpf = cpf_formatado
                
                except:
                    
                    time.sleep(0)    
            else:
                preparando_pagamento = False
                break

            if nome == "None":
                if valor == "None":
                    preparando_pagamento = False
                    break
            else:
                print("Estou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".")
            
            if isinstance(processo,str):
                processo = processo.replace('.','')
                processo = processo.replace('-','')
                processo = processo.replace('/','')
                processo_sem_pontos = int(processo)
            else: 
                print('Processo é inválido.')
            
            processo_formatado = re.sub(r'(\d{4})(\d{6})(\d{4})(\d{2})', r'\1.\2/\3-\4', "{:016d}".format(int(processo_sem_pontos))) 
            
            try:
                exercicio = int(exercicio)
            except:
                try:
                    empenho = int(empenho)
                except:
                    exercicio = 2025
                    exercicio = str(exercicio)
                    empenho = str(empenho)

            if isinstance(empenho,str):
                empenho = empenho.upper()
                exercicio = empenho.strip().split('NE')[0]
                empenho = empenho.strip().split('NE')[1]
                exercicio = int(exercicio)
                empenho = int(empenho)
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
            else: 
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
                
            if ja_foi_liquidado == True:
                    
                if ja_foi_preparado == False:

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                    campo_gestao = pp_despesa_empenhada.locator("#txtGestao_SIGEFPesquisa")
                    campo_gestao.wait_for(timeout=5000)
                    campo_gestao.dblclick()
                    campo_gestao.press_sequentially(gestao)
                    ponto_interrogacao = pp_despesa_empenhada.locator("#txtNotaLancamento_lnkBtnPesquisa")
                    
                    with pp_despesa_empenhada.expect_popup() as popup_info:
                        ponto_interrogacao.click()
                        time.sleep(0.5)
                        obedece_ou_nao_ordem_cronologica = popup_info.value
                        obedece_ou_nao_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                        obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Obedece Ordem Cronológica", exact=True)
                        nao_obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Não Obedece Ordem Cronológica")
                        nao_obedece.wait_for(timeout=5000)
                    
                        try:
                            with pp_despesa_empenhada.expect_popup() as popup_info:
                                nao_obedece.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica = popup_info.value
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                numero_nl = gerar_ordem_cronologica.locator("#txtNotaLancamento_SIGEFPesquisa")
                                numero_nl.wait_for(timeout=5000)
                                numero_exercicio = gerar_ordem_cronologica.locator('[name="txtNotaLancamentoSigla"]')
                                liquidacao = liquidacao.upper()
                                numero_liquidacao = liquidacao.strip().split('NL')[1]
                                exercicio_financeiro = liquidacao.strip().split('NL')[0]
                                numero_liquidacao_1 = int(numero_liquidacao)
                                exercicio_financeiro_1 = int(exercicio_financeiro)
                                exercicio_NL = str(exercicio_financeiro_1) + "NL"
                                nota_lancamento_formatada = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((numero_liquidacao_1)))
                                nota_lancamento_formatada = nota_lancamento_formatada.replace('0000NE',exercicio_NL)
                                numero_nl.press_sequentially(str(numero_liquidacao_1)) 
                                time.sleep(0.5) 
                                numero_exercicio.press_sequentially(str(exercicio_financeiro))
                                time.sleep(0.5) 
                                botao_confirmar = gerar_ordem_cronologica.get_by_role("button", name="Confirma a Consulta")
                                botao_confirmar.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                fonte_recurso = gerar_ordem_cronologica.locator('td[onclick="SelecionarItem(\'0\');"]')
                                fonte_recurso.wait_for()
                                fonte_recurso.click()
                                time.sleep(0.5)
                        except:
                            time.sleep(0)
                    try:
                            with pp_despesa_empenhada.expect_popup() as popup_info:
                                nao_obedece.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica = popup_info.value
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                numero_nl = gerar_ordem_cronologica.locator("#txtNotaLancamento_SIGEFPesquisa")
                                numero_nl.wait_for(timeout=5000)
                                numero_exercicio = gerar_ordem_cronologica.locator('[name="txtNotaLancamentoSigla"]')
                                liquidacao = liquidacao.upper()
                                numero_liquidacao = liquidacao.strip().split('NL')[1]
                                exercicio_financeiro = liquidacao.strip().split('NL')[0]
                                numero_liquidacao_1 = int(numero_liquidacao)
                                exercicio_financeiro_1 = int(exercicio_financeiro)
                                exercicio_NL = str(exercicio_financeiro_1) + "NL"
                                nota_lancamento_formatada = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((numero_liquidacao_1)))
                                nota_lancamento_formatada = nota_lancamento_formatada.replace('0000NE',exercicio_NL)
                                numero_nl.press_sequentially(str(numero_liquidacao_1)) 
                                time.sleep(0.5) 
                                numero_exercicio.press_sequentially(str(exercicio_financeiro))
                                time.sleep(0.5) 
                                botao_confirmar = gerar_ordem_cronologica.get_by_role("button", name="Confirma a Consulta")
                                botao_confirmar.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                fonte_recurso = gerar_ordem_cronologica.locator('td[onclick="SelecionarItem(\'0\');"]')
                                fonte_recurso.wait_for()
                                fonte_recurso.click()
                                time.sleep(0.5)
                    except:
                            time.sleep(0)

                    pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                    cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa")
                    cessionario.wait_for(timeout=5000)
                    value_cessionario = cessionario.input_value()
                    value_cessionario = value_cessionario
                    while value_cessionario != cpf_formatado:
                        cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa")
                        value_cessionario = cessionario.input_value()
                        value_cessionario = value_cessionario
                    else:
                        print('Liquidação encontrada.')
                    tipo_ordem_bancaria = pp_despesa_empenhada.locator("#cboTipoOrdemBancaria")
                    tipo_ordem_bancaria.wait_for(timeout=5000)
                    tipo_ordem_bancaria.select_option(label="Descentralizada")
                    #tipo_ordem_bancaria.select_option(label="Regularização")
                    locator_banco = pp_despesa_empenhada.locator("#txtBanco")
                    locator_banco.wait_for(timeout=5000)
                    locator_agencia = pp_despesa_empenhada.locator("#txtAgencia")
                    locator_conta_corrente = pp_despesa_empenhada.locator("#txtConta_SIGEFPesquisa")
                    
                    locator_banco.press_sequentially(banco)
                    time.sleep(0.5)
                    ponto_interrogacao2= pp_despesa_empenhada.locator("#txtConta_lnkBtnPesquisa")
                    ponto_interrogacao2.wait_for(timeout=5000)
                    with pp_despesa_empenhada.expect_popup() as popup_info:
                        
                        ponto_interrogacao2.click()
                        pesquisar_domicilio_bancario = popup_info.value
                        pesquisar_domicilio_bancario.wait_for_load_state('networkidle', timeout=30000)
                        botao_confirmar = pesquisar_domicilio_bancario.get_by_role("button", name="Confirmar a Consulta")
                        botao_confirmar.click()
                        time.sleep(0.5)
                        pesquisar_domicilio_bancario.wait_for_load_state('networkidle', timeout=30000)
                        try:
                            conta_nova = str(conta)
                            conta_nova = conta.replace("-",'')
                            conta_nova = conta.replace("-",'')
                            conta_nova = conta.upper()

                            conta_formatada_com_traco = re.sub(r'(.{9})(.{1})', r'\1-\2', "{:0>10}".format(conta_nova))
                            conta_formatada_sem_traco = conta_formatada_com_traco.replace("-",'')
                            
                            try:
                                selecionar_banco = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_sem_traco, exact=True)
                                selecionar_banco.wait_for(timeout=1000)
                            except:                                
                                selecionar_banco = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_com_traco, exact=True)
                                selecionar_banco.wait_for(timeout=1000)
                            
                            if selecionar_banco.is_visible():
                                conta_que_peguei = selecionar_banco.inner_text()
                                conta_que_peguei = conta_que_peguei.upper()
                        
                            print(f"Procurando pela célula da conta: '{conta}'...")
                            #pesquisar_domicilio_bancario.pause()
                          
                            #linha_correta = pesquisar_domicilio_bancario.locator("tr").filter(has_text=conta_formatada_sem_traco)
                            #linha_correta.wait_for(timeout=2000)
                            linha_correta = pesquisar_domicilio_bancario.locator("tr[class*='GridLinha']").filter(has_text=conta_formatada_sem_traco)
                        
                            #linha_correta = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_sem_traco, exact=True)
                            
                            print("Linha da conta encontrada na tabela.")

                            try:
                                conta_que_peguei = conta_que_peguei.replace("-",'')
                            except: 
                                time.sleep(0)
                        
                            try:
                                conta_que_peguei = conta_que_peguei.replace(".",'')
                            except: 
                                time.sleep(0)

                            if conta_que_peguei == conta_formatada_sem_traco:
                                time.sleep(0.5)
                                seletor_onclick = f'td[onclick*="{conta_formatada_sem_traco}"]'
                                try:
                                    celula_banco_para_clicar = linha_correta.first.get_by_role("cell", name=banco, exact=True)
                                    celula_banco_para_clicar.wait_for(timeout=3000)
                                    celula_banco_para_clicar.click()
                                except:
                                    celula_banco_para_clicar = linha_correta.get_by_role("cell", name=banco).nth(3)
                                    celula_banco_para_clicar.wait_for(timeout=3000)
                                    celula_banco_para_clicar.click()
                        except Exception as e:
                            print(f"Ocorreu um erro ao tentar selecionar a conta pela conta corrente: {e}")
                            selecione_manual = 'Selecione manualmente. A conta bancária inscrita na planilha é ' + banco + ' ' + agencia + ' ' +conta + '.'
                            pyautogui.alert(text=selecione_manual, title='Seleção Manual', button='OK')

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)        
                            

                    campo_observacao = pp_despesa_empenhada.locator("#txtObservacao")
                    campo_observacao.wait_for(timeout=5000)
                    texto_da_pp = "Preparação de Pagamento: Pagamento para o(a) servidor(a) " + str(cpf_formatado) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo_formatado) + "."
                    campo_observacao.press("Control+KeyA+Backspace")
                    campo_observacao.press_sequentially(texto_da_pp)
                    botao_retencoes = pp_despesa_empenhada.get_by_role("button", name="Sugerir Retenções")
                    botao_retencoes.wait_for(timeout=5000)
                    botao_retencoes.click()
                    sugerindo_retencoes = False
                    nao_existem_retencoes = pp_despesa_empenhada.get_by_text("Não existem sugestões para")
                    while sugerindo_retencoes == False:
                        if nao_existem_retencoes.is_visible():
                            sugerindo_retencoes=True
                    if sugerindo_retencoes == True:
                        time.sleep(0.5)
                        menu_confirmacao = pp_despesa_empenhada.locator("#menun7").get_by_role("link")
                        menu_confirmacao.click()
                        pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                        confirmacao_banco = pp_despesa_empenhada.locator("#txtBancoConf")
                        confirmacao_agencia = pp_despesa_empenhada.locator("#txtAgenciaConf")
                        confirmacao_conta = pp_despesa_empenhada.locator("#txtContaConf")
                        confirmacao_conta.wait_for(timeout=5000)
                        confirmacao_banco_value = confirmacao_banco.input_value()
                        confirmacao_agencia_value = confirmacao_agencia.input_value()
                        confirmacao_conta_value = confirmacao_conta.input_value()
                        value_confirmacao_conta = confirmacao_conta_value
                        value_confirmacao_agencia = confirmacao_agencia_value
                        value_confirmacao_banco = confirmacao_banco_value
                        value_confirmacao_conta = value_confirmacao_conta.upper()
                        
                        if robo_deve_parar:
                            verificar_panico_e_sair(book)

                        

                        if value_confirmacao_conta == conta_formatada_com_traco:

                            botao_confirmar = pp_despesa_empenhada.get_by_role("button", name="Confirmar a Operação")
                            botao_confirmar.wait_for(timeout=5000)
                            botao_confirmar.click()
                            try:
                                time.sleep(0.3)
                                mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                mensagem_sucesso.wait_for(timeout=5000)
                                texto_completo = mensagem_sucesso.inner_text()
                            except:
                                try:
                                    time.sleep(0.5)
                                    mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                    mensagem_sucesso.wait_for(timeout=5000)
                                    texto_completo = mensagem_sucesso.inner_text()
                                except:
                                    pp = 'erro'
                            
                            if "O número gerado foi" in texto_completo:
                                numero_nl = texto_completo.split("foi ")[1]
                                pp = numero_nl.strip('.')
                                print(pp)

                            botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                            botao_limpar.wait_for(timeout=5000)
                            botao_limpar.click()
                            pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                            
                            try:
                                pagina3_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp])
                                pagina3.delete_rows(linha,1)
                                pagina3.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp])
                                book.save(planilha)
                                pp_backup = pp
                                pp = 'não foi feita'
                            except:
                                book1.save("Pagamentos_Backup.xlsx")
                                print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                book1.close()
                                sys.exit()

                        else:
                            continuar = pyautogui.confirm(text='Domicílio Bancário diferente da planilha. Continuar?', title='Continuar' , buttons=['SIM', 'NÃO'])
                            continuar = str(continuar)
                            
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)

                            if continuar == 'SIM':
                                botao_confirmar = pp_despesa_empenhada.get_by_role("button", name="Confirmar a Operação")
                                botao_confirmar.click()
                                mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                mensagem_sucesso.wait_for(timeout=5000)
                                texto_completo = mensagem_sucesso.inner_text()
                                if "O número gerado foi" in texto_completo:
                                    numero_nl = texto_completo.split("foi ")[1]
                                    pp = numero_nl.strip('.')
                                    print(pp)
                                botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                                botao_limpar.wait_for(timeout=5000)
                                botao_limpar.click()
                                pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                                try:
                                    pagina3_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia,value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp])
                                    pagina3.delete_rows(linha,1)
                                    pagina3.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia,value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp])
                                    book.save(planilha)
                                    pp = 'não foi feita'
                                    
                                    if robo_deve_parar:
                                        verificar_panico_e_sair(book)
                                        try:
                                            pp_despesa_empenhada.close()
                                        except:
                                            time.sleep(0)
                                        
                                except:
                                    book1.save("Pagamentos_Backup.xlsx")
                                    print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                    book1.close()
                                    sys.exit()

                            else:
                                continuar = pyautogui.confirm(text='Deseja encerrar por aqui?', title='Continuar' , buttons=['SIM', 'NÃO'])
                                if continuar == 'SIM':

                                    pp = 'não foi feita'
                                    
                                    try:
                                        banco = "-"
                                        agencia ="-"
                                        conta = "-"
                                        pagina3_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia,conta,nota_de_empenho,despesa_certificada,liquidacao,pp])
                                        pagina3.delete_rows(linha,1)
                                        pagina3.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,banco, agencia, conta,nota_de_empenho,despesa_certificada,liquidacao,pp])
                                        book.save(planilha)
                                    
                                    except:
                                        
                                        book1.save("Pagamentos_Backup.xlsx")
                                        print("Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                        book1.close()
                                        sys.exit()
                                    
                                    botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                                    botao_limpar.wait_for(timeout=5000)
                                    botao_limpar.click()
                                else:
                                    if book:
                                        book.close()
                                        sys.exit()
                else:
                    linha = linha + 1
                    print("Próximo!")

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)
                             
print("Fim das preparações de Pagamento.")
if book:
    book.close()

try:
    pp_despesa_empenhada.close()
except:
    time.sleep(0)

print("\nScript finalizado. A janela de depuracao permanece aberta.") 
pyautogui.alert(text='Encerrei por aqui.', title='Fim', button='OK')
sys.exit()


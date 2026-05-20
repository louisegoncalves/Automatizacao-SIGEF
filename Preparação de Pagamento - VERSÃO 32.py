# =================================================================
#OLÁ!
#PROCEDIMENTO: PREPARAR PAGAMENTO;
#VERSÃO 32 EM 20/05/2026;
#POR: LOUISE-SESDEC;
#ALTERAÇÕES NO CÓDIGO PODEM SER ACESSADAS NO MEU GITHUB: <https://github.com/louisegoncalves/Automatizacao-SIGEF>.
# =================================================================

# =================================================================
#INSTRUÇÕES:
#ATENÇÃO: É OBRIGATÓRIO ABRIR O DEPURADOR DO GOOGLE CHROME PARA EXECUTAR ESSE CÓDIGO;
#EXECUTE NO CMD: ""C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebugProfile"
#E LOGUE NO SIGEF.
# =================================================================

# =================================================================
#BIBLIOTECAS UTILIZADAS:
import openpyxl
import pyautogui
import sys
import keyboard
from playwright.sync_api import sync_playwright
import time
import re
import os
from datetime import date
from datetime import datetime
# =================================================================

#QUAL PLANILHA VAI SER UTILIZADA?
planilha = "Pagamentos.xlsx"

#VARIÁVEIS IMPORTANTES
robo_deve_parar = False()
linha = 2
ainda_nao_foi_feito = '-'
existe_planilha = False
numero_de_operacoes = 0

#PLANILHAS NO EXCEL:
while existe_planilha == False:

    #PLANILHA PRINCIPAL:
    try:
        book = openpyxl.load_workbook(planilha)
        pagina1 = book['Entrada']
        pagina2 = book['Despesas Certificadas']
        pagina3 = book['Notas de Liquidação']
        pagina4 = book['Preparações de Pagamento']
        pagina5 = book['Ordens Bancárias']
        pagina6 = book['Saída']
        existe_planilha = True
    except: 
        pyautogui.alert(text='\nDeu algum erro na planilha.', title='Erro', button='OK')
        existe_planilha = False
        sys.exit()

    #PLANILHA DE BACKUP:
    try:
        book_backup = openpyxl.load_workbook("Backup.xlsx")
        pagina1_backup = book_backup['Entrada']
        pagina2_backup = book_backup['Despesas Certificadas']
        pagina3_backup = book_backup['Notas de Liquidação']
        pagina4_backup = book_backup['Preparações de Pagamento']
        pagina5_backup = book_backup['Ordens Bancárias']
        pagina6_backup = book_backup['Saída']
        existe_planilha = True
    except:
        existe_planilha = False
        try:
            wb = openpyxl.Workbook()
            ws_principal = wb.active
            ws_principal.title = "Entrada"
            wb.save("Backup.xlsx")
            try:
                wb.create_sheet("Despesas Certificadas")
                wb.create_sheet("Notas de Liquidação")
                wb.create_sheet("Preparações de Pagamento")
                wb.create_sheet("Ordens Bancárias")
                wb.create_sheet('Saída')
                wb.save("Backup.xlsx")
                print("\nArquivo 'Backup.xlsx' criado com sucesso com várias planilhas.")
            except:
                existe_planilha = False
                print("\nErro na planilha de backup.")
        except:
            existe_planilha = False
            print("\nErro na planilha de backup.")
    

# =================================================================
print('=================================================================')
print('PROCEDIMENTO: PREPARAR PAGAMENTO')
print('Versão 32 em 20/05/2026')
print('Por: LOUISE-SESDEC')
print('Github: https://github.com/louisegoncalves/Automatizacao-SIGEF')
print('=================================================================')
# =================================================================

#FUNÇÃO QUE SERÁ CHAMADA PELA TECLA DE PANICO
def parar_execucao():
    global robo_deve_parar
    print("\n!!! TECLA ESC ACIONADA! ENCERRANDO AUTOMACAO !!!")
    robo_deve_parar = True

#FUNÇÃO QUE ENCERRA O CODIGO E FECHA A PLANILHA COM SEGURANÇA
#A PLANILHA DEVE SEMPRE SER FECHADA ANTES DE ENCERRAR, POIS CORRE O RISCO DE CORROMPER
def verificar_panico_e_sair(workbook):
    global robo_deve_parar
    if robo_deve_parar:
        print("Garantindo que a planilha seja fechada...")
        if workbook:
            workbook.close()
        pyautogui.alert('Tecla ESC acionada. Automação encerrada.')
        sys.exit()

#DEFINA SUA TECLA DE PÂNICO
tecla_de_panico = "Esc" 
keyboard.add_hotkey(tecla_de_panico, parar_execucao)
print(f"Robô iniciado. Pressione a tecla '{tecla_de_panico}' a qualquer momento para abortar com seguranca.")

#PORTA DO DEPURADOR DO GOOGLE CHROME
CHROME_DEBUG_URL = "http://localhost:9222"

processo = str(pagina3.cell(row=linha,column=3).value)

#DEFININDO NÚMERO DE OPERAÇÕES:
while processo != "None":
    numero_de_operacoes = numero_de_operacoes + 1
    linha = linha + 1
    processo = str(pagina3.cell(row=linha,column=3).value)
else: 
    linha = 2
    linha_documento = linha - 1
    print("Planilha lida. São " + str(numero_de_operacoes) + " para executar.")

if robo_deve_parar:
    verificar_panico_e_sair(book)

#EXECUTANDO O PLAYWRIGHT DE FORMA SÍNCRONA
with sync_playwright() as p:
        if robo_deve_parar:
            verificar_panico_e_sair(book)

       #CONECTAR AO NAVEGADOR JÁ ABERTO:
        browser_conectado = False
        while browser_conectado == False:
            print(f"\nTentando se conectar ao Chrome na porta de depuração: {CHROME_DEBUG_URL}")
            try:
                browser = p.chromium.connect_over_cdp(CHROME_DEBUG_URL)
                print("\nConexão estabelecida com sucesso!")
                browser_conectado = True
            except: 
                #AQUI ELE VAI PEDIR PARA ABRIR O SIGEF PELO DEPURADOR DO GOOGLE
                pyautogui.confirm(text='Não identifiquei o SIGEF aberto. \n \nAperte OK quando o SIGEF estiver logado no depurador do Google Chrome', title='Depurador do Chrome' , buttons=['OK'])

        #OBTER A PÁGINA QUE ESTÁ ABERTA:
        janela = browser.contexts[0]
        guia = janela.pages[0]

        for p in janela.pages:
            if "SIGEF - Sistema Integrado de Planejamento e Gestão Fiscal" in p.title():
                guia = p
                break
            
        if guia is None:
            print("[ERRO]: Não encontrei nenhuma aba aberta com o SIGEF logado!")
        
        #VERIFICAR A PÁGINA ABERTA:
        print(f"Assumindo o controle da página com o título: '{guia.title()}'")
            
        #LOCALIZANDO O IFRAME:
        frame = guia.frame_locator('iframe[src="/SIGEF2026/SEG/#/SEGControleAcesso?p=1"]')
        
        if robo_deve_parar:
            verificar_panico_e_sair(book)
        

        #SE QUISER DESATIVAR AQUELA JANELA DO COMEÇO PODE EXCLUIR ELA AQUI:
        pyautogui.alert(text='Procedimento: Preparar Pagamento. \n \nVersão 32 em 20/05/2026. \n \nPor: LOUISE-SESDEC', title='Início', button='OK')
        # =================================================================

        os.system('cls' if os.name == 'nt' else 'clear')
                          
        print("\n[INÍCIO DAS PREPARAÇÕES DE PAGAMENTO] Iniciando as preparações de pagamento.")
        print("\nPlanilha lida. São " + str(numero_de_operacoes) + " operações para executar.")

        pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
        pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
        pesquisar_funcionalidades_sistema.press_sequentially("PP Despesa Empenhada")
        funcionalidade_sistema = frame.get_by_title("PP Despesa Empenhada")
        
        with guia.expect_popup() as popup_info:
            funcionalidade_sistema.click()
            pp_despesa_empenhada = popup_info.value
            pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)

        
        while linha_documento <= numero_de_operacoes:

            #OBTENDO A UNIDADE GESTORA:
            ug = str(pagina3.cell(row=linha, column=1).value)

           #OBTENDO A GESTÃO:
            gestao = str(pagina3.cell(row=linha, column=2).value)

            #LENDO O Nº DO PROCESSO
            processo = str(pagina3.cell(row=linha,column=3).value)
            
            #LENDO O NOME DO SERVIDOR
            nome = str(pagina3.cell(row=linha, column=4).value)

            #LENDO O CPF
            cpf = str(pagina3.cell(row=linha, column=5).value)
                    
            #LENDO O VALOR
            valor = str(pagina3.cell(row=linha, column=6).value)
            
            #LENDO O BANCO
            banco = str(pagina1.cell(row=linha, column=7).value).upper()  
            try:
                banco = banco.replace(" ",'')
            except:
                time.sleep(0)
            banco_backup = banco
                    
            bancos = {
                                "BRASIL": "001", "BB": "001", "01": "001", "1": "001", "077": "077", "97": "097",
                                "BANCO DO BRASIL": "001", "CAIXA": "104", "NUBANK": "260", "ITAÚ": "341",
                                "ITAU": "341", "INTER": "077", "77": "077", "BRADESCO": "237", "104": "104",
                                "PICPAY": "380", "SANTANDER": "033", "33": "033", "CEF": "104", "756": "756", "336": "336",
                                "CREDISIS JICRED": "097", "CREDISIS": "097", "JICRED": "097", "097": "097", "380": "380",
                                "SICOOB": "756", "BANCOOB": "756", "CREDSIS": "097", "NUBANL": "260", "260": "260",
                                "PAN - 623": "623", "PAN": "623", "PIC PAY": "380", "PICPPAY": "380", "623": "623",
                                "C6": "336", "C6 BANK": "336", "001": "001", "033": "033", "237": "237", "341": "341",                   
            }

            banco = bancos.get(banco, '001')

            #LENDO A AGENCIA
            agencia = str(pagina1.cell(row=linha, column=8).value)
            try:
                agencia = agencia.replace(" ",'')
            except:
                time.sleep(0)

            #LENDO A CONTA CORRENTE
            conta = str(pagina1.cell(row=linha, column=9).value)
            try:
                conta = conta.replace(" ",'')
            except:
                time.sleep(0)

            #LENDO A NOTA DE EMPENHO
            empenho = str(pagina3.cell(row=linha, column=10).value)

            #LENDO A DESPESA CERTIFICADA
            despesa_certificada = str(pagina2.cell(row=linha, column=11).value)

            #LENDO A NOTA DE LIQUIDAÇÃO
            liquidacao = str(pagina3.cell(row=linha, column=12).value)
            if liquidacao != "None":
                ja_foi_liquidado = True
            else:
                ja_foi_liquidado = False
            
            #LENDO A PREPARAÇÃO DE PAGAMENTO
            preparacao_pagamento = str(pagina4.cell(row=linha, column=13).value)
        
            if preparacao_pagamento != "None":
                ja_foi_preparado = True
            else:
                ja_foi_preparado = False

            #LENDO A ORDEM BANCÁRIA
            ordem_bancaria = str(pagina5.cell(row=linha, column=14).value)

            #LENDO A DATA
            data = str(pagina3.cell(row=linha, column=15).value)

            #LENDO A OPERAÇÃO
            operacao = str(pagina3.cell(row=linha, column=16).value)

            #LENDO A DATA QUE DEVERÁ SER REALIZADO O PAGAMENTO
            data_do_pagamento = str(pagina3.cell(row=linha, column=17).value)

            #OBTENDO O NÚMERO DO DOCUMENTO:
            value_numero_cortado = str(pagina3.cell(row=linha, column=19).value)     

            if robo_deve_parar:
                verificar_panico_e_sair(book)
                pp_despesa_empenhada.close()

            if ja_foi_liquidado == True:
   
                while ja_foi_preparado == False:

                    print("\nEstou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".")

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    with pp_despesa_empenhada.expect_popup() as popup_info:
                        
                        data_do_pagamento_preencher = pp_despesa_empenhada.locator("#txtDataReferencia_SIGEFData")
                        data_do_pagamento_preencher.dblclick()
                        data_do_pagamento_preencher.fill(data_do_pagamento)
                        campo_gestao = pp_despesa_empenhada.locator("#txtGestao_SIGEFPesquisa")
                        campo_gestao.wait_for(timeout=5000)
                        campo_gestao.dblclick()
                        campo_gestao.fill(gestao)
                        ponto_interrogacao = pp_despesa_empenhada.locator("#txtNotaLancamento_lnkBtnPesquisa")
                        ponto_interrogacao.click()
                        time.sleep(0.1)
                        obedece_ou_nao_ordem_cronologica = popup_info.value
                        obedece_ou_nao_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                        obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Obedece Ordem Cronológica", exact=True)
                        nao_obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Não Obedece Ordem Cronológica")
                        nao_obedece.wait_for(timeout=5000)
                    
                    try:
                        with pp_despesa_empenhada.expect_popup() as popup_info:
                                nao_obedece.click()
                                time.sleep(0.1)
                                gerar_ordem_cronologica = popup_info.value
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                numero_nl = gerar_ordem_cronologica.locator("#txtNotaLancamento_SIGEFPesquisa")
                                numero_nl.wait_for(timeout=5000)
                                numero_exercicio = gerar_ordem_cronologica.locator('[name="txtNotaLancamentoSigla"]')
                                liquidacao = liquidacao.upper()
                                numero_liquidacao = liquidacao.strip().split('NL')[1]
                                exercicio_financeiro = liquidacao.strip().split('NL')[0]
                                numero_liquidacao_1 = int(numero_liquidacao)
                                exercicio_financeiro_1 = int(exercicio_financeiro)
                                exercicio_NL = str(exercicio_financeiro_1) + "NL"
                                nota_lancamento_formatada = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((numero_liquidacao_1)))
                                nota_lancamento_formatada = nota_lancamento_formatada.replace('0000NE',exercicio_NL)
                                numero_nl.fill(str(numero_liquidacao_1)) 
                                numero_exercicio.fill(str(exercicio_financeiro))
                                botao_confirmar = gerar_ordem_cronologica.get_by_role("button", name="Confirma a Consulta")
                                botao_confirmar.click()
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                fonte_recurso = gerar_ordem_cronologica.locator('td[onclick="SelecionarItem(\'0\');"]')
                                fonte_recurso.wait_for()
                                fonte_recurso.click()
                    except:
                            time.sleep(0)
                            try:
                                with pp_despesa_empenhada.expect_popup() as popup_info:
                                    ponto_interrogacao.click()
                                    time.sleep(0.1)
                                    obedece_ou_nao_ordem_cronologica = popup_info.value
                                    obedece_ou_nao_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                    time.sleep(0.3)
                                    obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Obedece Ordem Cronológica", exact=True)
                                    nao_obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Não Obedece Ordem Cronológica")
                                    nao_obedece.wait_for(timeout=5000)
                                with pp_despesa_empenhada.expect_popup() as popup_info:
                                    nao_obedece.click()
                                    time.sleep(0.1)
                                    gerar_ordem_cronologica = popup_info.value
                                    gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                    numero_nl = gerar_ordem_cronologica.locator("#txtNotaLancamento_SIGEFPesquisa")
                                    numero_nl.wait_for(timeout=5000)
                                    numero_exercicio = gerar_ordem_cronologica.locator('[name="txtNotaLancamentoSigla"]')
                                    liquidacao = liquidacao.upper()
                                    numero_liquidacao = liquidacao.strip().split('NL')[1]
                                    exercicio_financeiro = liquidacao.strip().split('NL')[0]
                                    numero_liquidacao_1 = int(numero_liquidacao)
                                    exercicio_financeiro_1 = int(exercicio_financeiro)
                                    exercicio_NL = str(exercicio_financeiro_1) + "NL"
                                    nota_lancamento_formatada = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((numero_liquidacao_1)))
                                    nota_lancamento_formatada = nota_lancamento_formatada.replace('0000NE',exercicio_NL)
                                    numero_nl.fill(str(numero_liquidacao_1)) 
                                    numero_exercicio.fill(str(exercicio_financeiro))
                                    botao_confirmar = gerar_ordem_cronologica.get_by_role("button", name="Confirma a Consulta")
                                    botao_confirmar.click()
                                    gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                    fonte_recurso = gerar_ordem_cronologica.locator('td[onclick="SelecionarItem(\'0\');"]')
                                    fonte_recurso.wait_for()
                                    fonte_recurso.click()
                            except:
                                time.sleep(0)

                    pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                    cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa")
                    cessionario.wait_for(timeout=5000)
                    value_cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa").input_value()

                    while value_cessionario != cpf:
                        pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                        cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa")
                        cessionario.wait_for(timeout=5000)
                        value_cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa").input_value()
                        #print("[ATENÇÃO] Validando dados...")
                        if robo_deve_parar:
                            pp_despesa_empenhada.close()
                            verificar_panico_e_sair(book)       
                    else:
                        print(f"[VALIDAÇÃO] Liquidação: '{liquidacao}'")
                    
                    tipo_ordem_bancaria = pp_despesa_empenhada.locator("#cboTipoOrdemBancaria")
                    tipo_ordem_bancaria.wait_for(timeout=5000)
                    tipo_ordem_bancaria.select_option(label="Descentralizada")
                    #tipo_ordem_bancaria.select_option(label="Regularização")
                    locator_banco = pp_despesa_empenhada.locator("#txtBanco")
                    locator_banco.wait_for(timeout=5000)
                    locator_agencia = pp_despesa_empenhada.locator("#txtAgencia")
                    locator_conta_corrente = pp_despesa_empenhada.locator("#txtConta_SIGEFPesquisa")
                    
                    try:
                        natureza_despesa = pp_despesa_empenhada.locator("#txtNaturezaDespesa")
                        value_natureza_despesa = natureza_despesa.input_value()
       
                    except:
                        time.sleep(0)
                    
                    locator_banco.fill(banco)
                    ponto_interrogacao2= pp_despesa_empenhada.locator("#txtConta_lnkBtnPesquisa")
                    ponto_interrogacao2.wait_for(timeout=5000)
                    time.sleep(0.1)
                    with pp_despesa_empenhada.expect_popup() as popup_info:
                        ponto_interrogacao2.click()
                        pesquisar_domicilio_bancario = popup_info.value
                        pesquisar_domicilio_bancario.wait_for_load_state('networkidle', timeout=30000)
                        botao_confirmar = pesquisar_domicilio_bancario.get_by_role("button", name="Confirmar a Consulta")
                        botao_confirmar.click()
                        pesquisar_domicilio_bancario.wait_for_load_state('networkidle', timeout=30000)
                        try:
                            conta_nova = str(conta)
                            conta_nova = conta_nova.replace(",",'')
                            conta_nova = conta_nova.replace(".",'')
                            conta_nova = conta_nova.replace("-",'')
                            conta_nova = conta_nova.replace(" ",'')
                            conta_nova = conta_nova.upper()

                            conta_formatada_com_traco = re.sub(r'(.{9})(.{1})', r'\1-\2', "{:0>10}".format(conta_nova))
                            conta_formatada_sem_traco = conta_formatada_com_traco.replace("-",'')
                            
                            try:
                                selecionar_banco = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_sem_traco, exact=True)
                                selecionar_banco.wait_for(timeout=1000)
                            except:                                
                                selecionar_banco = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_com_traco, exact=True)
                                selecionar_banco.wait_for(timeout=1000)
                            
                            if selecionar_banco.is_visible():
                                conta_que_peguei = selecionar_banco.inner_text()
                                conta_que_peguei = conta_que_peguei.upper()
                        
                            print(f"[VALIDAÇÃO] Procurando pela célula da conta: '{conta}'...")
                            #pesquisar_domicilio_bancario.pause()
                            #linha_correta = pesquisar_domicilio_bancario.locator("tr").filter(has_text=conta_formatada_sem_traco)
                            #linha_correta.wait_for(timeout=2000)
                            linha_correta = pesquisar_domicilio_bancario.locator("tr[class*='GridLinha']").filter(has_text=conta_formatada_sem_traco)
                            #linha_correta = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_sem_traco, exact=True)

                            print("[VALIDAÇÃO] Linha da conta encontrada na tabela.")

                            try:
                                conta_que_peguei = conta_que_peguei.replace("-",'')
                            except: 
                                time.sleep(0)
                        
                            try:
                                conta_que_peguei = conta_que_peguei.replace(".",'')
                            except: 
                                time.sleep(0)
                            
                            try:
                                conta_formatada_sem_traco = conta_formatada_sem_traco.upper()
                            except: 
                                time.sleep(0)

                            if conta_que_peguei == conta_formatada_sem_traco:
                                seletor_onclick = f'td[onclick*="{conta_formatada_sem_traco}"]'
                                try:
                                    celula_banco_para_clicar = linha_correta.first.get_by_role("cell", name=banco, exact=True)
                                    celula_banco_para_clicar.wait_for(timeout=1000)
                                    celula_banco_para_clicar.click()
                                except:
                                    celula_banco_para_clicar = linha_correta.get_by_role("cell", name=banco).nth(3)
                                    celula_banco_para_clicar.wait_for(timeout=1000)
                                    celula_banco_para_clicar.click()
                        except Exception as e:
                            print(f"[SELECIONE MANUAL] Ocorreu um erro ao tentar selecionar a conta pela conta corrente: {e}")
                            selecione_manual = 'Selecione manualmente. A conta bancária inscrita na planilha é ' + banco_backup + ' ' + agencia + ' ' +conta + '.'
                            pyautogui.alert(text=selecione_manual, title='Seleção Manual', button='OK')

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)
                        pp_despesa_empenhada.close()        
                            
                    #INFORMAÇÕES PRELIMINARES
                    #HORA:
                    agora = datetime.now()

                    campo_observacao = pp_despesa_empenhada.locator("#txtObservacao")
                    campo_observacao.wait_for(timeout=5000)
                    texto_da_pp = "Preparação de Pagamento: Pagamento para o(a) servidor(a) " + str(cpf) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo) + "."
                    campo_observacao.press("Control+KeyA+Backspace")
                    campo_observacao.fill(texto_da_pp)
                    botao_retencoes = pp_despesa_empenhada.get_by_role("button", name="Sugerir Retenções")
                    botao_retencoes.wait_for(timeout=5000)
                    botao_retencoes.click()
                    sugerindo_retencoes = False
                    nao_existem_retencoes = pp_despesa_empenhada.get_by_text("Não existem sugestões para")
                    
                    while sugerindo_retencoes == False:
                        if nao_existem_retencoes.is_visible():
                            sugerindo_retencoes=True
                    if sugerindo_retencoes == True:

                        if value_natureza_despesa == "33.90.93.01":
                            botao_reinf = pp_despesa_empenhada.locator("#menun6").get_by_role("link")
                            botao_reinf.click()
                            botao_interrogacao_reinf = pp_despesa_empenhada.locator("#txtNaturezaRendimento_lnkBtnPesquisa")

                            with pp_despesa_empenhada.expect_popup() as popup_info:
                                botao_interrogacao_reinf.click()

                                janela_reinf = popup_info.value
                                janela_reinf.wait_for_load_state('networkidle', timeout=30000)

                                botao_confirmar = janela_reinf.get_by_role("button", name="Confirmar a Consulta")
                                botao_confirmar.wait_for(timeout=3000)
                                botao_confirmar.click()

                                codigo_10001 = janela_reinf.get_by_role("cell", name="10001", exact=True)
                                codigo_10001.wait_for(timeout=3000)
                                codigo_10001.click()
                        
                        
                        
                        menu_confirmacao = pp_despesa_empenhada.locator("#menun7").get_by_role("link")
                        time.sleep(0.1)
                        menu_confirmacao.click()
                        pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                        confirmacao_banco = pp_despesa_empenhada.locator("#txtBancoConf")
                        confirmacao_agencia = pp_despesa_empenhada.locator("#txtAgenciaConf")
                        confirmacao_conta = pp_despesa_empenhada.locator("#txtContaConf")
                        confirmacao_conta.wait_for(timeout=5000)
                        confirmacao_banco_value = confirmacao_banco.input_value()
                        confirmacao_agencia_value = confirmacao_agencia.input_value()
                        confirmacao_conta_value = confirmacao_conta.input_value()
                        value_confirmacao_conta = confirmacao_conta_value
                        value_confirmacao_agencia = confirmacao_agencia_value
                        value_confirmacao_banco = confirmacao_banco_value
                        value_confirmacao_conta = value_confirmacao_conta.upper()
   
                        if robo_deve_parar:
                            verificar_panico_e_sair(book)
                            pp_despesa_empenhada.close()

                        conta_formatada_com_traco = conta_formatada_com_traco.upper()
                        
                        if value_confirmacao_conta == conta_formatada_com_traco:
                            botao_confirmar = pp_despesa_empenhada.get_by_role("button", name="Confirmar a Operação")
                            botao_confirmar.wait_for(timeout=2000)
                            botao_confirmar.click()
                            
                            try:
                                time.sleep(0.3)
                                mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                mensagem_sucesso.wait_for(timeout=5000)
                                texto_completo = mensagem_sucesso.inner_text()
                            except:
                                try:
                                    time.sleep(0.3)
                                    mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                    mensagem_sucesso.wait_for(timeout=5000)
                                    texto_completo = mensagem_sucesso.inner_text()
                                except:
                                    pp = 'erro'
                            
                            if "O número gerado foi" in texto_completo:
                                numero_nl = texto_completo.split("foi ")[1]
                                pp = numero_nl.strip('.')
                                print(f"[SUCESSO] Preparação de Pagamento encontrada e copiada: '{pp}'")
                                ja_foi_preparado = True

                            botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                            botao_limpar.wait_for(timeout=2000)
                            botao_limpar.click()
                            pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                                pp_despesa_empenhada.close()
                            
                            try:
                                print("Inserindo dados na planilha...")

                                pagina4_backup.append([ug,gestao,processo,nome,cpf,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado])
                                book_backup.save("Backup.xlsx")

                                                                
                                dados = [ug,gestao,processo,nome,cpf,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado]

                                for numero_coluna, valor in enumerate(dados, start=1):
                                    pagina4.cell(row=linha, column=numero_coluna, value=valor)

                                book.save(planilha)
                                print("[SUCESSO] Planilha salva.")
                                
                            except:
                                book_backup.save("Backup.xlsx")
                                print("[ERRO NA PLANILHA] Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                book_backup.close()
                                sys.exit()

                        else:
                            continuar = pyautogui.confirm(text='Domicílio Bancário diferente da planilha. Continuar?', title='Continuar' , buttons=['SIM', 'NÃO'])
                            continuar = str(continuar)
                            
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                                pp_despesa_empenhada.close()

                            if continuar == 'SIM':
                                botao_confirmar = pp_despesa_empenhada.get_by_role("button", name="Confirmar a Operação")
                                botao_confirmar.click()
                                mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                mensagem_sucesso.wait_for(timeout=5000)
                                texto_completo = mensagem_sucesso.inner_text()
                                if "O número gerado foi" in texto_completo:
                                    numero_nl = texto_completo.split("foi ")[1]
                                    pp = numero_nl.strip('.')
                                    print(f"[SUCESSO] Preparação de Pagamento encontrada e copiada: '{pp}'")
                                botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                                botao_limpar.wait_for(timeout=5000)
                                botao_limpar.click()
                                pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                                try:
                                    pagina4_backup.append([ug,gestao,processo,nome,cpf,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado])
                                    
                                    book_backup.save("Backup.xlsx")
                                    
                                    if robo_deve_parar:
                                        verificar_panico_e_sair(book)
                                        pp_despesa_empenhada.close()
                                except:
                                    book_backup.save("Backup.xlsx")
                                    print("[ERRO NA PLANILHA] Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                    book_backup.close()
                                    sys.exit()

                            else:
                                continuar = pyautogui.confirm(text='Deseja encerrar por aqui?', title='Continuar' , buttons=['SIM', 'NÃO'])
                                
                                if continuar == 'NÃO':
                                    

                                    pp = "Servidor foi pulado, PP não foi feita."

                                    try:
                                        banco = "-"
                                        agencia = "-"
                                        conta = "-"
                                        pagina4_backup.append([ug,gestao,processo,nome,cpf,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado])
                                        book_backup.save("Backup.xlsx")

                                                                
                                        dados = [ug,gestao,processo,nome,cpf,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_do_pagamento,agora,value_numero_cortado]

                                        for numero_coluna, valor in enumerate(dados, start=1):
                                            pagina4.cell(row=linha, column=numero_coluna, value=valor)

                                        book.save(planilha)
                                    except:
                                        
                                        book_backup.save("Backup.xlsx")
                                        print("[ERRO NA PLANILHA] Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                        book_backup.close()
                                        sys.exit()
                                    
                                    botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                                    botao_limpar.wait_for(timeout=2000)
                                    botao_limpar.click()
                                else:
                                    if book:
                                        book.close()
                                        sys.exit()
                else:
                    linha = linha + 1
                    linha_documento = linha - 1
                    pp = 'não foi feita'

                    if robo_deve_parar:
                        pp_despesa_empenhada.close()
                        verificar_panico_e_sair(book)
        else:
            print("\n[FIM DAS PREPARAÇÕES DE PAGAMENTO] Nenhuma despesa para preparar.")

if book:
    book.close()
print("\nScript finalizado. A janela de depuração permanece aberta.") 
pyautogui.alert(text='Encerrei por aqui.', title='Fim', button='OK')
sys.exit()
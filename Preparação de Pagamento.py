#OLÁ!
#PROCEDIMENTO: PREPARAR PAGAMENTO;
#POR: LOUISE-SESDEC;
#ALTERAÇÕES NO CÓDIGO PODEM SER ACESSADAS NO MEU GITHUB: <https://github.com/louisegoncalves/Automatizacao-SIGEF>.

#INSTRUÇÕES
#ATENÇÃO: É OBRIGATÓRIO ABRIR O DEPURADOR DO GOOGLE CHROME PARA EXECUTAR ESSE CÓDIGO
#EXECUTE NO CMD: ""C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebugProfile"
#E LOGUE NO SIGEF

#BIBLIOTECAS UTILIZADAS:
import openpyxl
import pyautogui
import sys
import keyboard
from playwright.sync_api import sync_playwright
import time
import re
from datetime import date
from datetime import datetime

#QUAL PLANILHA VAI SER UTILIZADA?
planilha = "Pagamentos - voluntariar dezembro.xlsx"
#planilha = "Pagamentos.xlsx"

#VARIÁVEIS IMPORTANTES
robo_deve_parar = False
coluna = 1
linha = 2
preparando_pagamento = True
ainda_nao_foi_feito = '-'

#PLANILHA NO EXCEL:
try:
    book = openpyxl.load_workbook(planilha)
    pagina = book['Entrada']
    pagina1 = book['Despesas Certificadas']
    pagina2 = book['Notas de Liquidação']
    pagina3 = book['Preparações de Pagamento']
    pagina4 = book['Ordens Bancárias']
    pagina5 = book['Saída']
except: 
    pyautogui.alert(text='Deu algum erro na planilha.', title='Erro', button='OK')
    sys.exit()      
try:
    book1 = openpyxl.load_workbook("Pagamentos_Backup.xlsx")
    pagina_backup = book1['Entrada']
    pagina1_backup = book1['Despesas Certificadas']
    pagina2_backup = book1['Notas de Liquidação']
    pagina3_backup = book1['Preparações de Pagamento']
    pagina4_backup = book1['Ordens Bancárias']
    pagina5_backup = book1['Saída']
except:
    try:
        wb = openpyxl.Workbook()
        ws_principal = wb.active
        ws_principal.title = "Entrada"
        wb.save("Pagamentos_Backup.xlsx")
        try:
            wb.create_sheet("Despesas Certificadas")
            wb.create_sheet("Notas de Liquidação")
            wb.create_sheet("Preparações de Pagamento")
            wb.create_sheet("Ordens Bancárias")
            wb.create_sheet('Saída')
            wb.save("Pagamentos_Backup.xlsx")
            print("Arquivo 'Pagamentos_Backup.xlsx' criado com sucesso com várias planilhas.")
        except:
            print("Planilha de backup encontrada.")
        try:
            book1 = openpyxl.load_workbook("Pagamentos_Backup.xlsx")
            pagina_backup = book1['Entrada']
            pagina1_backup = book1['Despesas Certificadas']
            pagina2_backup = book1['Notas de Liquidação']
            pagina3_backup = book1['Preparações de Pagamento']
            pagina4_backup = book1['Ordens Bancárias']
            pagina5_backup = book1['Saída']
        except:
            print("Erro na planilha de backup.")
    except:
        print("Erro na planilha de backup.")
         
#SE QUISER DESATIVAR AQUELA JANELA DO COMEÇO PODE EXCLUIR ELA AQUI:
pyautogui.alert(text='Procedimento: PP Despesa Empenhada.', title='Início', button='OK')

#FUNÇÃO QUE SERÁ CHAMADA PELA TECLA DE PANICO
def parar_execucao():
    global robo_deve_parar
    print("\n!!! TECLA ESC ACIONADA! ENCERRANDO AUTOMACAO !!!")
    robo_deve_parar = True

#FUNÇÃO QUE ENCERRA O CODIGO E FECHA A PLANILHA COM SEGURANÇA
#A PLANILHA DEVE SEMPRE SER FECHADA ANTES DE ENCERRAR, POIS CORRE O RISCO DE CORROMPER
def verificar_panico_e_sair(workbook):
    global robo_deve_parar
    if robo_deve_parar:
        print("Garantindo que a planilha seja fechada...")
        if workbook:
            workbook.close()
        pyautogui.alert('Tecla ESC acionada. Automação encerrada.')
        sys.exit()

#DEFINA SUA TECLA DE PÂNICO
tecla_de_panico = "Esc" 
keyboard.add_hotkey(tecla_de_panico, parar_execucao)
print(f"Robô iniciado. Pressione a tecla '{tecla_de_panico}' a qualquer momento para abortar com seguranca.")

#AQUI ELE VAI PEDIR PARA ABRIR O SIGEF PELO DEPURADOR DO GOOGLE
pyautogui.confirm(text='Aperte OK quando o SIGEF estiver logado no depurador do Google Chrome', title='Depurador do Chrome' , buttons=['OK'])

#PORTA DO DEPURADOR DO GOOGLE CHROME
CHROME_DEBUG_URL = "http://localhost:9222"

if robo_deve_parar:
    verificar_panico_e_sair(book)

#EXECUTANDO O PLAYWRIGHT DE FORMA SÍNCRONA
with sync_playwright() as p:
        if robo_deve_parar:
            verificar_panico_e_sair(book)

        #CONECTAR AO NAVEGADOR JÁ ABERTO:
        print(f"Tentando se conectar ao Chrome na porta de depuração: {CHROME_DEBUG_URL}")
        browser = p.chromium.connect_over_cdp(CHROME_DEBUG_URL)
        print("Conexão estabelecida com sucesso!")

        #OBTER A PÁGINA QUE ESTÁ ABERTA:
        #Quando conectamos, precisamos pegar o contexto e a página existentes;
        #Geralmente, a primeira página do primeiro contexto é a que queremos.
        janela = browser.contexts[0]
        guia = janela.pages[0]

        #VERIFICAR A PÁGINA ABERTA:
        print(f"Assumindo o controle da página com o título: '{guia.title()}'")
            
        #LOCALIZANDO O IFRAME:
        frame = guia.frame_locator('iframe[src="/SIGEF2025/SEG/#/SEGControleAcesso?p=1"]')
        
        if robo_deve_parar:
            verificar_panico_e_sair(book)
        
        #INÍCIO                
        print("\nIniciando!")
        pesquisar_funcionalidades_sistema = frame.get_by_placeholder("Pesquisar funcionalidades do sistema...")
        pesquisar_funcionalidades_sistema.press("Control+KeyA+Backspace")
        pesquisar_funcionalidades_sistema.press_sequentially("PP Despesa Empenhada")
        funcionalidade_sistema = frame.get_by_title("PP Despesa Empenhada")
        
        with guia.expect_popup() as popup_info:
            funcionalidade_sistema.click()
            pp_despesa_empenhada = popup_info.value
        
        linha = 2

        while preparando_pagamento == True:
            
            coluna = 1

            #LENDO A UG
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ug = cell.value
                    ug= str(ug)

            #LENDO A GESTÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    gestao = cell.value
                    gestao = str(gestao)

            #LENDO O Nº DO PROCESSO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    processo = cell.value
                    processo = str(processo)
            
            #LENDO O NOME DO SERVIDOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    nome = cell.value
                    nome = str(nome)
                    primeiro_nome = nome.split()[0]

            #LENDO O CPF
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    cpf = cell.value
                    cpf = str(cpf)
                    
            #LENDO O VALOR
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):                
                for cell in row:
                    valor = cell.value
                    valor = str(valor)
                    try:
                        valor = valor.replace('R$','')
                    except:
                        time.sleep(0)
                    try:
                        valor = valor.replace(' ','')
                    except:
                        time.sleep(0)
                    try:
                        valor = valor.replace('.','')
                    except:
                        time.sleep(0)
            
            #LENDO O BANCO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    banco = cell.value
                    banco = str(banco)
                    
            banco_backup = banco
            try:
                banco = banco.replace(' ','')
            except:
                time.sleep(0)
            try:
                    banco_formatado = re.sub(r'(\d{2})(\d{1})', r'\1-\2', "{:03d}".format(int(banco)))
                    banco_formatado = banco_formatado.replace("-","")
                    deu_certo_a_formatacao_do_banco = True
                    banco_deduzido = False
            except:
                    deu_certo_a_formatacao_do_banco = False
                    time.sleep(0)
                    
            if deu_certo_a_formatacao_do_banco == False:
                        banco = banco.upper()
                        if "BRASIL" in banco:
                            banco = '001'
                            banco_deduzido = True
                            deu_certo_a_formatacao_do_banco = False
                        if "NEXT" in banco:
                            banco = '237'
                            banco_deduzido = True
                            deu_certo_a_formatacao_do_banco = False
                        if "BRADESCO" in banco:
                            banco = '237'
                            banco_deduzido = True
                            deu_certo_a_formatacao_do_banco = False
                        if "NUBANK" in banco:
                            banco = '260' 
                            banco_deduzido = True
                            deu_certo_a_formatacao_do_banco = False
                        if "NÚBANK" in banco:
                            banco = '260'   
                            banco_deduzido = True  
                            deu_certo_a_formatacao_do_banco = False                    
                        if "COOB" in banco:
                            banco = '756'    
                            banco_deduzido = True   
                            deu_certo_a_formatacao_do_banco = False      
                        if "SICOOB" in banco:
                            banco = '756'    
                            banco_deduzido = True   
                            deu_certo_a_formatacao_do_banco = False       
                        if "BANCOOB" in banco:
                            banco = '756'    
                            banco_deduzido = True   
                            deu_certo_a_formatacao_do_banco = False          
                        if "CAIXA" in banco:
                            banco = '104'    
                            banco_deduzido = True    
                            deu_certo_a_formatacao_do_banco = False                
                        if "CEF" in banco:
                            banco = '104'  
                            banco_deduzido = True    
                            deu_certo_a_formatacao_do_banco = False                  
                        if "ITA" in banco:
                            banco = '341'  
                            banco_deduzido = True    
                            deu_certo_a_formatacao_do_banco = False                  
                        if "INTER" in banco:
                            banco = '077'  
                            banco_deduzido = True 
                            deu_certo_a_formatacao_do_banco = False                     
                        if "BB" in banco:
                            banco = '001'  
                            banco_deduzido = True     
                            deu_certo_a_formatacao_do_banco = False                 
                        if "PIC" in banco:
                            banco = '380'  
                            banco_deduzido = True    
                            deu_certo_a_formatacao_do_banco = False     
                        if "PICPAY" in banco:
                            banco = '380'  
                            banco_deduzido = True    
                            deu_certo_a_formatacao_do_banco = False                
                        if "C6" in banco:
                            banco = '336'  
                            banco_deduzido = True  
                            deu_certo_a_formatacao_do_banco = False                    
                        if "CRED" in banco:
                            banco = '097'   
                            banco_deduzido = True    
                            deu_certo_a_formatacao_do_banco = False        
                        if "CREDISIS" in banco:
                            banco = '097'   
                            banco_deduzido = True    
                            deu_certo_a_formatacao_do_banco = False          
                        if "SANTANDER" in banco:
                            banco = '033'
                            banco_deduzido = True
                            deu_certo_a_formatacao_do_banco = False
                        if "PAN" in banco:
                            banco = '623'
                            banco_deduzido = True
                            deu_certo_a_formatacao_do_banco = False
            else:
                        banco = banco_formatado
                        
            if banco_deduzido == False:
                        banco = '001'

            #LENDO A AGENCIA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    agencia = cell.value
                    agencia = str(agencia)

            #LENDO A CONTA CORRENTE
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    conta = cell.value
                    conta = str(conta)

            #LENDO A NOTA DE EMPENHO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    empenho = cell.value
                    empenho = str(empenho)

            #LENDO A DESPESA CERTIFICADA
            coluna = coluna + 1
            for row in pagina1.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    despesa_certificada = cell.value
                    despesa_certificada = str(despesa_certificada)

            #LENDO A NOTA DE LIQUIDAÇÃO
            coluna = coluna + 1
            for row in pagina2.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    liquidacao = cell.value
                    liquidacao = str(liquidacao)
            
            #LENDO A PREPARAÇÃO DE PAGAMENTO
            coluna = coluna + 1
            for row in pagina3.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    preparacao_pagamento = cell.value
                    preparacao_pagamento = str(preparacao_pagamento)

            #LENDO A ORDEM BANCÁRIA
            coluna = coluna + 1
            for row in pagina4.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    ordem_bancaria = cell.value
                    ordem_bancaria = str(ordem_bancaria)

            #LENDO A DATA
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    data = cell.value
                    data = str(data)

            #LENDO A OPERAÇÃO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    operacao = cell.value
                    operacao = str(operacao)

            #LENDO A DATA QUE DEVERÁ SER REALIZADO O PAGAMENTO
            coluna = coluna + 1
            for row in pagina.iter_rows(min_row=linha,max_col=coluna,max_row=linha):
                for cell in row:
                    data_do_pagamento = cell.value
                    data_do_pagamento = str(data_do_pagamento)
            

            if data != "None":
                if isinstance(data_do_pagamento,str):
                    data_formatada = data_do_pagamento.replace('/','')
                    data_formatada = data_formatada.replace('.','')
                    data_formatada = data_formatada.replace('-','')
                    data_formatada = data_formatada.replace(' ','')
                    data_formatada = data_formatada.replace(',','')
                            
                    try:
                        data_formatada = int(data_formatada)
                        data_formatada = re.sub(r'(\d{2})(\d{2})(\d{4})', r'\1/\2/\3', "{:08d}".format(int(data_formatada)))
                        data_foi_formatada = True
                    except:
                        data_atual = date.today() 
                        data_formatada = data_atual.strftime("%d/%m/%Y")
                        data_foi_formatada = True
                        
                else:
                    data_atual = date.today() 
                    data_formatada = data_atual.strftime("%d/%m/%Y")
                    data_foi_formatada = True
                        
            else: 
                print('[ATENÇÃO] Data não foi preenchida na planilha.')
                data_foi_formatada = False
            
            if liquidacao != None:
                ja_foi_liquidado = True
                
                if preparacao_pagamento != "None":
                    ja_foi_preparado = True
                else:
                    ja_foi_preparado = False
            else:
                ja_foi_liquidado = False
                preparando_pagamento = False

            if robo_deve_parar:
                verificar_panico_e_sair(book)
            
            if cpf != "None":
                try:
                    if isinstance(cpf,str):
                        cpf = cpf.replace('.','')
                        cpf = cpf.replace('-','')
                        cpf_sem_ponto_virgula = int(cpf)
                    else: 
                        print('[ATENÇÃO] CPF é inválido.')
                    
                    cpf_formatado = re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', "{:011d}".format(int(cpf_sem_ponto_virgula)))                        
                    cpf = cpf_formatado
                
                except:
                    time.sleep(0)
                        
            else:
                preparando_pagamento = False
                break

            if nome == "None":
                if valor == "None":
                    preparando_pagamento = False
                    break
            else:
                time.sleep(0)
            
            if isinstance(processo,str):

                processo = processo.replace('.','')
                processo = processo.replace('-','')
                processo = processo.replace('/','')
                processo_sem_pontos = int(processo)
            else: 
                print('[ATENÇÃO] Processo é inválido.')
            
            processo_formatado = re.sub(r'(\d{4})(\d{6})(\d{4})(\d{2})', r'\1.\2/\3-\4', "{:016d}".format(int(processo_sem_pontos))) 

            try:
                exercicio = int(exercicio)
            except:
                try:
                    empenho = int(empenho)
                except:
                    exercicio = 2025
                    exercicio = str(exercicio)
                    empenho = str(empenho)

            if isinstance(empenho,str):
                empenho = empenho.upper()
                exercicio = empenho.strip().split('NE')[0]
                empenho = empenho.strip().split('NE')[1]
                exercicio = int(exercicio)
                empenho = int(empenho)
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
            else: 
                exercicio_NE = str(exercicio) + "NE"
                nota_de_empenho = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((empenho)))
                nota_de_empenho = nota_de_empenho.replace('0000NE',exercicio_NE)
                
            
            if ja_foi_liquidado == True:
   
                if ja_foi_preparado == False:

                    print("\nEstou na linha " + str(linha) + " da planilha, referente ao servidor " + str(nome) + ".")

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)

                    pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                    data_do_pagamento_preencher = pp_despesa_empenhada.locator("#txtDataReferencia_SIGEFData")
                    data_do_pagamento_preencher.dblclick()
                    data_do_pagamento_preencher.press_sequentially(data_formatada)
                    campo_gestao = pp_despesa_empenhada.locator("#txtGestao_SIGEFPesquisa")
                    campo_gestao.wait_for(timeout=5000)
                    campo_gestao.dblclick()
                    campo_gestao.press_sequentially(gestao)
                    ponto_interrogacao = pp_despesa_empenhada.locator("#txtNotaLancamento_lnkBtnPesquisa")
                    
                    with pp_despesa_empenhada.expect_popup() as popup_info:
                        ponto_interrogacao.click()
                        time.sleep(0.5)
                        obedece_ou_nao_ordem_cronologica = popup_info.value
                        obedece_ou_nao_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                        obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Obedece Ordem Cronológica", exact=True)
                        nao_obedece = obedece_ou_nao_ordem_cronologica.get_by_text("Não Obedece Ordem Cronológica")
                        nao_obedece.wait_for(timeout=5000)
                    
                        try:
                            with pp_despesa_empenhada.expect_popup() as popup_info:
                                nao_obedece.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica = popup_info.value
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                numero_nl = gerar_ordem_cronologica.locator("#txtNotaLancamento_SIGEFPesquisa")
                                numero_nl.wait_for(timeout=5000)
                                numero_exercicio = gerar_ordem_cronologica.locator('[name="txtNotaLancamentoSigla"]')
                                liquidacao = liquidacao.upper()
                                numero_liquidacao = liquidacao.strip().split('NL')[1]
                                exercicio_financeiro = liquidacao.strip().split('NL')[0]
                                numero_liquidacao_1 = int(numero_liquidacao)
                                exercicio_financeiro_1 = int(exercicio_financeiro)
                                exercicio_NL = str(exercicio_financeiro_1) + "NL"
                                nota_lancamento_formatada = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((numero_liquidacao_1)))
                                nota_lancamento_formatada = nota_lancamento_formatada.replace('0000NE',exercicio_NL)
                                numero_nl.press_sequentially(str(numero_liquidacao_1)) 
                                time.sleep(0.5) 
                                numero_exercicio.press_sequentially(str(exercicio_financeiro))
                                time.sleep(0.5) 
                                botao_confirmar = gerar_ordem_cronologica.get_by_role("button", name="Confirma a Consulta")
                                botao_confirmar.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                fonte_recurso = gerar_ordem_cronologica.locator('td[onclick="SelecionarItem(\'0\');"]')
                                fonte_recurso.wait_for()
                                fonte_recurso.click()
                                time.sleep(0.5)
                        except:
                            time.sleep(0)
                    try:
                            with pp_despesa_empenhada.expect_popup() as popup_info:
                                nao_obedece.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica = popup_info.value
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                numero_nl = gerar_ordem_cronologica.locator("#txtNotaLancamento_SIGEFPesquisa")
                                numero_nl.wait_for(timeout=5000)
                                numero_exercicio = gerar_ordem_cronologica.locator('[name="txtNotaLancamentoSigla"]')
                                liquidacao = liquidacao.upper()
                                numero_liquidacao = liquidacao.strip().split('NL')[1]
                                exercicio_financeiro = liquidacao.strip().split('NL')[0]
                                numero_liquidacao_1 = int(numero_liquidacao)
                                exercicio_financeiro_1 = int(exercicio_financeiro)
                                exercicio_NL = str(exercicio_financeiro_1) + "NL"
                                nota_lancamento_formatada = re.sub(r'(\d{4})(\d{6})', r'\1NE\2', "{:010d}".format((numero_liquidacao_1)))
                                nota_lancamento_formatada = nota_lancamento_formatada.replace('0000NE',exercicio_NL)
                                numero_nl.press_sequentially(str(numero_liquidacao_1)) 
                                time.sleep(0.5) 
                                numero_exercicio.press_sequentially(str(exercicio_financeiro))
                                time.sleep(0.5) 
                                botao_confirmar = gerar_ordem_cronologica.get_by_role("button", name="Confirma a Consulta")
                                botao_confirmar.click()
                                time.sleep(0.5)
                                gerar_ordem_cronologica.wait_for_load_state('networkidle', timeout=30000)
                                fonte_recurso = gerar_ordem_cronologica.locator('td[onclick="SelecionarItem(\'0\');"]')
                                fonte_recurso.wait_for()
                                fonte_recurso.click()
                                time.sleep(0.5)
                    except:
                            time.sleep(0)

                    pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                    cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa")
                    cessionario.wait_for(timeout=5000)
                    value_cessionario = cessionario.input_value()
                    value_cessionario = value_cessionario
                    while value_cessionario != cpf_formatado:
                        cessionario = pp_despesa_empenhada.locator("#txtCredor_SIGEFPesquisa")
                        value_cessionario = cessionario.input_value()
                        value_cessionario = value_cessionario
                    else:
                        print('[VALIDAÇÃO] Liquidação encontrada.')
                    tipo_ordem_bancaria = pp_despesa_empenhada.locator("#cboTipoOrdemBancaria")
                    tipo_ordem_bancaria.wait_for(timeout=5000)
                    tipo_ordem_bancaria.select_option(label="Descentralizada")
                    #tipo_ordem_bancaria.select_option(label="Regularização")
                    locator_banco = pp_despesa_empenhada.locator("#txtBanco")
                    locator_banco.wait_for(timeout=5000)
                    locator_agencia = pp_despesa_empenhada.locator("#txtAgencia")
                    locator_conta_corrente = pp_despesa_empenhada.locator("#txtConta_SIGEFPesquisa")
                    
                    locator_banco.press_sequentially(banco)
                    time.sleep(0.5)
                    ponto_interrogacao2= pp_despesa_empenhada.locator("#txtConta_lnkBtnPesquisa")
                    ponto_interrogacao2.wait_for(timeout=5000)
                    with pp_despesa_empenhada.expect_popup() as popup_info:
                        
                        ponto_interrogacao2.click()
                        pesquisar_domicilio_bancario = popup_info.value
                        pesquisar_domicilio_bancario.wait_for_load_state('networkidle', timeout=30000)
                        botao_confirmar = pesquisar_domicilio_bancario.get_by_role("button", name="Confirmar a Consulta")
                        botao_confirmar.click()
                        time.sleep(0.5)
                        pesquisar_domicilio_bancario.wait_for_load_state('networkidle', timeout=30000)
                        try:
                            conta_nova = str(conta)
                            conta_nova = conta.replace("-",'')
                            conta_nova = conta.replace("-",'')
                            conta_nova = conta.upper()

                            conta_formatada_com_traco = re.sub(r'(.{9})(.{1})', r'\1-\2', "{:0>10}".format(conta_nova))
                            conta_formatada_sem_traco = conta_formatada_com_traco.replace("-",'')
                            
                            try:
                                selecionar_banco = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_sem_traco, exact=True)
                                selecionar_banco.wait_for(timeout=1000)
                            except:                                
                                selecionar_banco = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_com_traco, exact=True)
                                selecionar_banco.wait_for(timeout=1000)
                            
                            if selecionar_banco.is_visible():
                                conta_que_peguei = selecionar_banco.inner_text()
                                conta_que_peguei = conta_que_peguei.upper()
                        
                            print(f"[VALIDAÇÃO] Procurando pela célula da conta: '{conta}'...")
                            #pesquisar_domicilio_bancario.pause()
                          
                            #linha_correta = pesquisar_domicilio_bancario.locator("tr").filter(has_text=conta_formatada_sem_traco)
                            #linha_correta.wait_for(timeout=2000)
                            linha_correta = pesquisar_domicilio_bancario.locator("tr[class*='GridLinha']").filter(has_text=conta_formatada_sem_traco)
                        
                            #linha_correta = pesquisar_domicilio_bancario.get_by_role("cell", name=conta_formatada_sem_traco, exact=True)
                            
                            print("[VALIDAÇÃO] Linha da conta encontrada na tabela.")

                            try:
                                conta_que_peguei = conta_que_peguei.replace("-",'')
                            except: 
                                time.sleep(0)
                        
                            try:
                                conta_que_peguei = conta_que_peguei.replace(".",'')
                            except: 
                                time.sleep(0)

                            if conta_que_peguei == conta_formatada_sem_traco:
                                time.sleep(0.5)
                                seletor_onclick = f'td[onclick*="{conta_formatada_sem_traco}"]'
                                try:
                                    celula_banco_para_clicar = linha_correta.first.get_by_role("cell", name=banco, exact=True)
                                    celula_banco_para_clicar.wait_for(timeout=1000)
                                    celula_banco_para_clicar.click()
                                except:
                                    celula_banco_para_clicar = linha_correta.get_by_role("cell", name=banco).nth(3)
                                    celula_banco_para_clicar.wait_for(timeout=1000)
                                    celula_banco_para_clicar.click()
                        except Exception as e:
                            print(f"[SELECIONE MANUAL] Ocorreu um erro ao tentar selecionar a conta pela conta corrente: {e}")
                            selecione_manual = 'Selecione manualmente. A conta bancária inscrita na planilha é ' + banco_backup + ' ' + agencia + ' ' +conta + '.'
                            pyautogui.alert(text=selecione_manual, title='Seleção Manual', button='OK')

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)        
                            
                    #INFORMAÇÕES PRELIMINARES
                    #HORA:
                    agora = datetime.now()

                    campo_observacao = pp_despesa_empenhada.locator("#txtObservacao")
                    campo_observacao.wait_for(timeout=5000)
                    texto_da_pp = "Preparação de Pagamento: Pagamento para o(a) servidor(a) " + str(cpf_formatado) + " " + str(nome) + " referente à " + str(operacao) + " realizada no período de " + str(data) + ". Processo Administrativo n: " + str(processo_formatado) + "."
                    campo_observacao.press("Control+KeyA+Backspace")
                    campo_observacao.press_sequentially(texto_da_pp)

                    botao_retencoes = pp_despesa_empenhada.get_by_role("button", name="Sugerir Retenções")
                    botao_retencoes.wait_for(timeout=5000)
                    botao_retencoes.click()
                    sugerindo_retencoes = False
                    nao_existem_retencoes = pp_despesa_empenhada.get_by_text("Não existem sugestões para")
                    while sugerindo_retencoes == False:
                        if nao_existem_retencoes.is_visible():
                            sugerindo_retencoes=True
                    if sugerindo_retencoes == True:
                        time.sleep(0.5)
                        menu_confirmacao = pp_despesa_empenhada.locator("#menun7").get_by_role("link")
                        menu_confirmacao.click()
                        pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                        confirmacao_banco = pp_despesa_empenhada.locator("#txtBancoConf")
                        confirmacao_agencia = pp_despesa_empenhada.locator("#txtAgenciaConf")
                        confirmacao_conta = pp_despesa_empenhada.locator("#txtContaConf")
                        confirmacao_conta.wait_for(timeout=5000)
                        confirmacao_banco_value = confirmacao_banco.input_value()
                        confirmacao_agencia_value = confirmacao_agencia.input_value()
                        confirmacao_conta_value = confirmacao_conta.input_value()
                        value_confirmacao_conta = confirmacao_conta_value
                        value_confirmacao_agencia = confirmacao_agencia_value
                        value_confirmacao_banco = confirmacao_banco_value
                        value_confirmacao_conta = value_confirmacao_conta.upper()
                        
                        if robo_deve_parar:
                            verificar_panico_e_sair(book)
                        
                        if value_confirmacao_conta == conta_formatada_com_traco:

                            botao_confirmar = pp_despesa_empenhada.get_by_role("button", name="Confirmar a Operação")
                            botao_confirmar.wait_for(timeout=5000)
                            botao_confirmar.click()
                            try:
                                time.sleep(0.3)
                                mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                mensagem_sucesso.wait_for(timeout=5000)
                                texto_completo = mensagem_sucesso.inner_text()
                            except:
                                try:
                                    time.sleep(0.5)
                                    mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                    mensagem_sucesso.wait_for(timeout=5000)
                                    texto_completo = mensagem_sucesso.inner_text()
                                except:
                                    pp = 'erro'
                            
                            if "O número gerado foi" in texto_completo:
                                numero_nl = texto_completo.split("foi ")[1]
                                pp = numero_nl.strip('.')
                                print(f"[SUCESSO] Preparação de Pagamento encontrada e copiada: '{pp}'")


                            botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                            botao_limpar.wait_for(timeout=5000)
                            botao_limpar.click()
                            pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)
                            
                            try:
                                pagina3_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_formatada,agora])
                                pagina3.delete_rows(linha,1)
                                pagina3.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_formatada,agora])

                                book.save(planilha)
                                pp_backup = pp
                                pp = 'não foi feita'
                            except:
                                book1.save("Pagamentos_Backup.xlsx")
                                print("[ERRO NA PLANILHA] Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                book1.close()
                                sys.exit()

                        else:
                            continuar = pyautogui.confirm(text='Domicílio Bancário diferente da planilha. Continuar?', title='Continuar' , buttons=['SIM', 'NÃO'])
                            continuar = str(continuar)
                            
                            if robo_deve_parar:
                                verificar_panico_e_sair(book)

                            if continuar == 'SIM':
                                botao_confirmar = pp_despesa_empenhada.get_by_role("button", name="Confirmar a Operação")
                                botao_confirmar.click()
                                mensagem_sucesso = pp_despesa_empenhada.get_by_text("Operação realizada com")
                                mensagem_sucesso.wait_for(timeout=5000)
                                texto_completo = mensagem_sucesso.inner_text()
                                if "O número gerado foi" in texto_completo:
                                    numero_nl = texto_completo.split("foi ")[1]
                                    pp = numero_nl.strip('.')
                                    print(f"[SUCESSO] Preparação de Pagamento encontrada e copiada: '{pp}'")
                                botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                                botao_limpar.wait_for(timeout=5000)
                                botao_limpar.click()
                                pp_despesa_empenhada.wait_for_load_state('networkidle', timeout=30000)
                                try:
                                    pagina3_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_formatada,agora])
                                    pagina3.delete_rows(linha,1)
                                    pagina3.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_formatada,agora])
                                    book.save(planilha)
                                    pp = 'não foi feita'
                                    if robo_deve_parar:
                                        verificar_panico_e_sair(book)
                                except:
                                    book1.save("Pagamentos_Backup.xlsx")
                                    print("[ERRO NA PLANILHA] Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                    book1.close()
                                    sys.exit()

                            else:
                                continuar = pyautogui.confirm(text='Deseja encerrar por aqui?', title='Continuar' , buttons=['SIM', 'NÃO'])
                                
                                if continuar == 'NÃO':

                                    pp = 'não foi feita'
                                    
                                    try:
                                        banco = "-"
                                        agencia ="-"
                                        conta = "-"
                                        pagina3_backup.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_formatada,agora])
                                        pagina3.delete_rows(linha,1)
                                        pagina3.append([ug,gestao,processo_formatado,nome,cpf_formatado,valor,value_confirmacao_banco, value_confirmacao_agencia, value_confirmacao_conta,nota_de_empenho,despesa_certificada,liquidacao,pp,ainda_nao_foi_feito,data,operacao,data_formatada,agora])
                                        book.save(planilha)
                                    
                                    except:
                                        
                                        book1.save("Pagamentos_Backup.xlsx")
                                        print("[ERRO NA PLANILHA] Deu algum erro ao salvar a planilha, a planilha de backup foi solicitada.")
                                        book1.close()
                                        sys.exit()
                                    
                                    botao_limpar = pp_despesa_empenhada.get_by_role("link", name="Limpar a Tela")
                                    botao_limpar.wait_for(timeout=5000)
                                    botao_limpar.click()
                                else:
                                    if book:
                                        book.close()
                                        sys.exit()
                else:
                    linha = linha + 1
                    banco_deduzido = False
                    deu_certo_a_formatacao_do_banco = False

                    if robo_deve_parar:
                        verificar_panico_e_sair(book)
                             
print("\nFim das preparações de Pagamento.")
if book:
    book.close()
print("\nScript finalizado. A janela de depuração permanece aberta.") 
pyautogui.alert(text='Encerrei por aqui.', title='Fim', button='OK')
sys.exit()
